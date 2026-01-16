import openpyxl
from openpyxl import Workbook
import tkinter as tk
from tkinter import filedialog
import sys

# --- CONFIGURACIÓN ---
# Cambia este código si tu rojo es diferente (FFFF0000 es el rojo estándar puro)
CODIGO_COLOR_ROJO = 'FFFF0000' 

def seleccionar_archivo():
    # Inicializar tkinter y ocultar la ventana principal pequeña
    root = tk.Tk()
    root.withdraw()

    print("Por favor, selecciona tu archivo Excel en la ventana emergente...")
    
    # Abrir el explorador de archivos
    ruta_archivo = filedialog.askopenfilename(
        title="Selecciona el archivo Excel a procesar",
        filetypes=[("Archivos de Excel", "*.xlsx")]
    )
    
    return ruta_archivo

# --- INICIO DEL PROCESO ---

# 1. Pedir al usuario que seleccione el archivo
ruta = seleccionar_archivo()

if not ruta:
    print("No seleccionaste ningún archivo. El programa se detendrá.")
    sys.exit()

print(f"Procesando el archivo: {ruta}")

# 2. Cargar el Excel seleccionado
try:
    wb_origen = openpyxl.load_workbook(ruta, data_only=True) # data_only lee valores, no fórmulas
    ws_origen = wb_origen.active
except Exception as e:
    print(f"Error al abrir el archivo: {e}")
    sys.exit()

# 3. Preparar los libros de salida
wb_rojos = Workbook()
ws_rojos = wb_rojos.active
ws_rojos.title = "Datos Rojos"

wb_blancos = Workbook()
ws_blancos = wb_blancos.active
ws_blancos.title = "Datos Limpios"

# 4. Iterar y separar
contador_rojos = 0
contador_blancos = 0

print("Analizando colores...")

for fila in ws_origen.iter_rows():
    # Obtenemos la celda A de cada fila para verificar el color
    # (Asumiendo que si la fila es roja, la primera celda también lo es)
    celda_verificacion = fila[0] 
    
    # Obtenemos los valores de toda la fila para copiarlos
    valores_fila = [celda.value for celda in fila]
    
    # Obtener el color (fgColor.rgb)
    # A veces el color es None si no tiene formato
    color = celda_verificacion.fill.fgColor.rgb if celda_verificacion.fill.fgColor else "00000000"

    # Comparar
    if color == CODIGO_COLOR_ROJO:
        ws_rojos.append(valores_fila)
        contador_rojos += 1
    else:
        ws_blancos.append(valores_fila)
        contador_blancos += 1

# 5. Guardar resultados
wb_rojos.save('Resultado_Rojos.xlsx')
wb_blancos.save('Resultado_Limpios.xlsx')

print("--- PROCESO TERMINADO ---")
print(f"Se encontraron {contador_rojos} filas ROJAS.")
print(f"Se encontraron {contador_blancos} filas LIMPIAS/BLANCAS.")
print("Se han creado dos archivos nuevos en esta misma carpeta.")