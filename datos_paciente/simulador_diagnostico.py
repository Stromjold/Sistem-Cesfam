import pandas as pd
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
import tkinter as tk
from tkinter import filedialog
import platform
import subprocess
import time

# Intentar importar funciones de generador_info si está disponible
try:
    from generador_info import generar_rut, format_rut
except ImportError:
    # Versión simplificada si no se puede importar
    def generar_rut():
        numero = random.randint(5000000, 25000000)
        rut_str = str(numero)
        suma = 0
        multiplo = 2
        for i in range(len(rut_str) - 1, -1, -1):
            suma += int(rut_str[i]) * multiplo
            multiplo = 2 if multiplo == 7 else multiplo + 1
        dv = 11 - (suma % 11)
        digito_verificador = '0' if dv == 11 else 'K' if dv == 10 else str(dv)
        return f"{rut_str[:-3]}.{rut_str[-3:]}-{digito_verificador}"

def print_header():
    """Imprime encabezado del programa"""
    print("\n" + "="*70)
    print(" "*10 + "🏥 GENERADOR DE DIAGNÓSTICOS MÉDICOS (CESFAM)")
    print("="*70 + "\n")

def clear_screen():
    """Limpia la pantalla según el SO - DESHABILITADO POR SOLICITUD"""
    pass 

def abrir_archivo_xlsx(ruta_archivo: str):
    """Abre un archivo xlsx con la aplicación predeterminada del sistema"""
    try:
        if os.name == 'nt':  # Windows
            os.startfile(ruta_archivo)
        elif os.name == 'posix':  # macOS/Linux
            sistema = platform.system()
            if sistema == 'Darwin':  # macOS
                subprocess.call(['open', ruta_archivo])
            else:  # Linux
                subprocess.call(['xdg-open', ruta_archivo])
        print("\nAbriendo archivo...")
    except OSError as e:
        print(f"\nNo se pudo abrir el archivo automáticamente: {e}")
        print(f"   Por favor, ábrelo manualmente desde: {ruta_archivo}")

# Lista de diagnósticos CIE-10 abreviada para Atención Primaria
DIAGNOSTICOS_CIE10 = [
    {'codigo': 'I10', 'descripcion': 'Hipertensión esencial (primaria)', 'categoria': 'Crónico'},
    {'codigo': 'E11', 'descripcion': 'Diabetes mellitus tipo 2', 'categoria': 'Crónico'},
    {'codigo': 'J00', 'descripcion': 'Rinofaringitis aguda (resfriado común)', 'categoria': 'Agudo'},
    {'codigo': 'J20', 'descripcion': 'Bronquitis aguda', 'categoria': 'Agudo'},
    {'codigo': 'K29', 'descripcion': 'Gastritis y duodenitis', 'categoria': 'Agudo'},
    {'codigo': 'M54.5', 'descripcion': 'Lumbago no especificado', 'categoria': 'Dolor'},
    {'codigo': 'R51', 'descripcion': 'Cefalea', 'categoria': 'Dolor'},
    {'codigo': 'N39.0', 'descripcion': 'Infección de vías urinarias, sitio no especificado', 'categoria': 'Infeccioso'},
    {'codigo': 'Z00.0', 'descripcion': 'Examen médico general', 'categoria': 'Preventivo'},
    {'codigo': 'F32', 'descripcion': 'Episodio depresivo', 'categoria': 'Salud Mental'},
    {'codigo': 'F41', 'descripcion': 'Otros trastornos de ansiedad', 'categoria': 'Salud Mental'},
    {'codigo': 'L20', 'descripcion': 'Dermatitis atópica', 'categoria': 'Dermatológico'},
    {'codigo': 'J45', 'descripcion': 'Asma', 'categoria': 'Crónico Respiratorio'},
    {'codigo': 'E66', 'descripcion': 'Obesidad', 'categoria': 'Nutricional'}
]

MEDICOS_TRATANTES = [
    'Dr. Juan Pérez', 'Dra. María González', 'Dr. Roberto Soto', 'Dra. Ana Silva', 
    'Dr. Carlos Muñoz', 'Dra. Paula Rojas', 'Enf. Luisa Tapia'
]

def seleccionar_archivo_input(titulo_ventana="Seleccionar archivo de pacientes existente"):
    """Abre diálogo para seleccionar archivo de pacientes, similar a separador_datos"""
    try:
        root = tk.Tk()
        root.withdraw()
        root.attributes('-topmost', True)
        
        filename = filedialog.askopenfilename(
            title=titulo_ventana,
            filetypes=[("Archivos Excel y CSV", "*.xlsx *.xls *.csv"), ("Todos los archivos", "*.*")]
        )
        
        root.destroy()
        return filename
    except Exception as e:
        print(f"❌ Error abriendo diálogo: {e}")
        return None

def obtener_ruts_desde_archivo(ruta):
    """Extrae lista de RUTs de un archivo Excel"""
    try:
        print(f"  📖 Leyendo archivo: {os.path.basename(ruta)}...")
        df = pd.read_excel(ruta)
        # Buscar columna de RUT
        cols_posibles = ['RUT', 'rut', 'Rut', 'RUN', 'run', 'Run', 'ID', 'id', 'Id_rut']
        col_rut = next((col for col in df.columns if col in cols_posibles), None)
        
        if col_rut:
            # Filtrar vacíos y convertir a string
            ruts = df[col_rut].dropna().astype(str).tolist()
            print(f"  ✓ Columna de RUT detectada: '{col_rut}' ({len(ruts)} registros)")
            return ruts
        else:
            print("  ⚠ No se encontró una columna de RUT válida (RUT, Run, ID)")
            return None
    except Exception as e:
        print(f"  ❌ Error leyendo archivo: {e}")
        return None

def generar_fecha_diagnostico(inicio_anio=2024):
    """Genera una fecha aleatoria reciente"""
    inicio = datetime(inicio_anio, 1, 1)
    dias_totales = (datetime.now() - inicio).days
    fecha_random = inicio + timedelta(days=random.randint(0, dias_totales))
    return fecha_random.strftime("%d/%m/%Y")

def generar_diagnosticos(cantidad, ruts_base=None):
    """Genera una lista de diagnósticos para pacientes simulados"""
    registros = []
    
    for i in range(cantidad):
        diag = random.choice(DIAGNOSTICOS_CIE10)
        
        # Simular estado del diagnóstico
        estado = random.choice(['Confirmado', 'Confirmado', 'Confirmado', 'Sospecha', 'Descartado'])
        
        if diag['categoria'] == 'Crónico':
            tratamiento = random.choice(['En tratamiento farmacológico', 'Control dieta y ejercicio', 'Sin adherencia', 'Control periódico'])
        elif diag['categoria'] == 'Agudo':
            tratamiento = random.choice(['Reposo y líquidos', 'Antibióticos', 'Sintomático', 'Derivación'])
        else:
            tratamiento = 'En evaluación'
            
        # Determinar RUT: de la lista base o generado
        rut_paciente = random.choice(ruts_base) if ruts_base else generar_rut()

        registro = {
            'ID_Atencion': 10000 + i,
            'Fecha': generar_fecha_diagnostico(),
            'RUT_Paciente': rut_paciente,
            'Codigo_CIE10': diag['codigo'],
            'Diagnostico': diag['descripcion'],
            'Categoria': diag['categoria'],
            'Estado': estado,
            'Medico_Tratante': random.choice(MEDICOS_TRATANTES),
            'Observaciones': tratamiento,
            'Requiere_Control': 'Sí' if random.random() > 0.6 else 'No'
        }
        registros.append(registro)
        
    return registros

def formato_excel_diagnostico(df: pd.DataFrame, ruta_archivo: str):
    """Guarda y da formato visual al Excel de diagnósticos"""
    try:
        df.to_excel(ruta_archivo, index=False, engine='openpyxl')
        
        wb = openpyxl.load_workbook(ruta_archivo)
        ws = wb.active
        
        # Validar que exista una hoja activa
        if ws is None:
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                print("⚠ Error: El archivo Excel no tiene hojas visibles.")
                return

        # Estilos
        header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid") # Azul médico
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Aplicar formato
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
        
        # Ajustar anchos
        for col_num, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws.column_dimensions[get_column_letter(col_num)].width = min(length + 2, 60)
            
        wb.save(ruta_archivo)
        print(f"✓ Archivo guardado: {ruta_archivo}")
        
    except Exception as e:
        print(f"⚠ Error guardando Excel: {e}")

def main():
    clear_screen()
    print_header()
    
    print("Este programa genera diagnósticos médicos simulados (CIE-10).")
    print("Puede generar RUTs aleatorios o usar una lista de pacientes existente.\n")

    modo = input("¿Desea cargar pacientes desde un archivo existente? (s/n): ").lower().strip()
    ruts_pacientes = None
    
    if modo == 's':
        print("\nAbriendo ventana de selección...")
        ruta_archivo = seleccionar_archivo_input()
        if ruta_archivo:
            ruts_pacientes = obtener_ruts_desde_archivo(ruta_archivo)
            if not ruts_pacientes:
                print("  ⚠ Se procederá con generación de RUTs aleatorios.")
        else:
            print("  ⚠ No se seleccionó archivo. Se generarán RUTs nuevos.")

    try:
        cant_str = input("\nIngrese cantidad de diagnósticos a generar (Enter para 50): ").strip()
        cantidad = int(cant_str) if cant_str else 50
    except ValueError:
        cantidad = 50
        print("  ⚠ Valor inválido, usando 50 por defecto.")
        
    nombre_archivo = input("Nombre del archivo de salida (Enter para 'reporte_diagnosticos.xlsx'): ").strip() or "reporte_diagnosticos.xlsx"
    
    if not nombre_archivo.endswith('.xlsx'):
        nombre_archivo += '.xlsx'
        
    print(f"\n⚙️ Generando {cantidad} registros de diagnóstico...")
    datos = generar_diagnosticos(cantidad, ruts_base=ruts_pacientes)
    df = pd.DataFrame(datos)
    
    formato_excel_diagnostico(df, nombre_archivo)
    
    print("\n📊 Resumen de diagnósticos generados:")
    print(df['Diagnostico'].value_counts().head().to_string())
    
    print("\n" + "="*70)
    abrir_archivo = input("¿Desea abrir el archivo generado? (s/n): ").lower().strip()
    if abrir_archivo == 's':
        abrir_archivo_xlsx(nombre_archivo)

if __name__ == "__main__":
    main()
