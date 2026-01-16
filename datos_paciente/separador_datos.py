"""
COMPARADOR DE ARCHIVOS XLSX/CSV - OPTIMIZADO PARA GRANDES VOLÚMENES

Características:
- Procesamiento eficiente de archivos con más de 100,000 registros
- Optimizado para archivos grandes (>8MB, >8,000KB)
- Análisis de todas las hojas en archivos XLSX
- Uso de operaciones vectorizadas para mejor performance
- Categorización automática de columnas para optimizar memoria
- Procesamiento por chunks para archivos CSV muy grandes
- Lectura solo-datos (read_only) para reducir uso de memoria
- Uso de sets para comparaciones rápidas
- Escritura optimizada con xlsxwriter para archivos grandes

Versión optimizada para grandes volúmenes de datos (>8MB)
"""

# Imports de la biblioteca estándar
import os
import platform
import subprocess
import sys
import time
import tkinter as tk
from tkinter import filedialog
from typing import Optional, Tuple
import warnings

# Imports de terceros
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd

warnings.filterwarnings('ignore', category=DeprecationWarning)

COMMON_KEY_NAMES = ['id_rut',
                    'rut',
                    'RUT',
                    'id',
                    'id_usuario',
                    'usuario_id',
                    'ID', 
                    'documento', 
                    'doc', 
                    'cedula', 
                    'ficha', 
                    'folio', 
                    'caso', 
                    'n_solicitud', 
                    'identificador']

# Constantes para detección de nombres
NAME_COL_VARIANTS = [
    'nombres', 'nombre', 'nombres paciente', 'nombre paciente',
    'previsión nombres', 'nombres_beneficiario'
]
PATERNO_COL_VARIANTS = [
    'apellido paterno', 'paterno', 'apellidopaterno',
    'primer apellido', 'apellido 1', 'apellido_paterno'
]
MATERNO_COL_VARIANTS = [
    'apellido materno', 'materno', 'apellidomaterno',
    'segundo apellido', 'apellido 2', 'apellido_materno'
]
FULLNAME_COL_VARIANTS = [
    'nombre completo', 'nombre y apellido', 'apellidos y nombres',
    'nombre_completo', 'paciente', 'nombre beneficiario'
]

def get_col_by_variants(df: pd.DataFrame, variants: list) -> Optional[str]:
    """Busca una columna en el DataFrame que coincida con alguna de las variantes."""
    df_cols_lower = [c.lower().strip() for c in df.columns]
    for variant in variants:
        if variant in df_cols_lower:
            idx = df_cols_lower.index(variant)
            return df.columns[idx]
        # Búsqueda parcial segura
        for col_idx, col_name in enumerate(df_cols_lower):
            is_exact_match = variant == col_name
            is_partial_match = len(variant) > 4 and variant in col_name
            if is_exact_match or is_partial_match:
                return df.columns[col_idx]
    return None


def get_essential_columns(df: pd.DataFrame) -> list:
    """
    Identifica y retorna solo las columnas esenciales del DataFrame:
    RUT, Nombre, Apellido Paterno, Apellido Materno, y columnas de estado
    """
    essential_cols = []
    
    # Buscar columna de RUT
    rut_col = get_col_by_variants(df, COMMON_KEY_NAMES)
    if rut_col and rut_col in df.columns:
        essential_cols.append(rut_col)
    
    # Buscar columna de Nombre
    nombre_col = get_col_by_variants(df, NAME_COL_VARIANTS)
    if nombre_col and nombre_col in df.columns:
        essential_cols.append(nombre_col)
    
    # Buscar columna de Apellido Paterno
    paterno_col = get_col_by_variants(df, PATERNO_COL_VARIANTS)
    if paterno_col and paterno_col in df.columns:
        essential_cols.append(paterno_col)
    
    # Buscar columna de Apellido Materno
    materno_col = get_col_by_variants(df, MATERNO_COL_VARIANTS)
    if materno_col and materno_col in df.columns:
        essential_cols.append(materno_col)
    
    # Agregar columnas de estado si existen
    for col in df.columns:
        if col.startswith('ESTADO_EN_') or col == '__HOJA_ORIGEN__':
            essential_cols.append(col)
    
    return essential_cols


def clean_string_for_key(series: pd.Series) -> pd.Series:
    """
    Limpia strings para uso en claves: elimina espacios múltiples, whitespace, convierte a mayúsculas.
    """
    # Primero eliminar .0 final de floats convertidos a string
    s = series.astype(str).str.replace(r'\.0$', '', regex=True)
    
    return (s
            .str.replace(r'\s+', ' ', regex=True)  # Elimina espacios múltiples
            .str.strip()  # Elimina espacios al inicio y final
            .str.upper()  # Convierte a mayúsculas
            .replace('NAN', '')  # Reemplaza NaN strings
            .replace('NONE', '')  # Reemplaza None strings
            .str.replace('.', '', regex=False)  # Elimina puntos (separadores de miles)
            .replace('-', ''))  # Elimina guiones


def generate_custom_key(df: pd.DataFrame, df_name: str, fields: list) -> Tuple[Optional[pd.Series], str, list]:
    """
    Genera una clave personalizada basada en los campos seleccionados.
    fields: lista de campos ['rut', 'nombre', 'paterno', 'materno']
    """
    cols_found = []
    col_names = []
    key_parts = []

    fields_str = ', '.join(fields)
    print(f"  ⏳ {df_name}: Construyendo clave personalizada con: {fields_str}...")

    # Mapeo de campos a variantes
    field_variants = {
        'rut': COMMON_KEY_NAMES,
        'nombre': NAME_COL_VARIANTS,
        'paterno': PATERNO_COL_VARIANTS,
        'materno': MATERNO_COL_VARIANTS
    }

    for field in fields:
        if field in field_variants:
            variants = field_variants[field]
            col = get_col_by_variants(df, variants)

            if col:
                cols_found.append(col)
                key_part = clean_string_for_key(df[col])
                key_parts.append(key_part)
                # Nombre para descripción
                if field == 'paterno': col_names.append("Paterno")
                elif field == 'materno': col_names.append("Materno")
                elif field == 'nombre': col_names.append("Nombre")
                elif field == 'rut': col_names.append("RUT")
            else:
                print(f"    ⚠️ No se encontró columna para campo '{field}' en {df_name}")

    if not cols_found:
        return None, "", []

    # Combinar partes
    if len(key_parts) == 1:
        key_series = key_parts[0]
    else:
        # Usar pipe como separador
        key_series = key_parts[0]
        for part in key_parts[1:]:
            key_series = key_series + "|" + part

    desc = " + ".join(col_names)
    print(f"    ✓ {df_name}: Usando {desc}")
    return key_series, desc, cols_found

def generate_person_key(df: pd.DataFrame, df_name: str) -> Tuple[Optional[pd.Series], str, list]:
    """
    Intenta generar una clave única basada en Nombre + Apellidos.
    Retorna: (Serie con la clave, Mensaje descriptivo, Lista de columnas usadas)
    """
    col_nombre = get_col_by_variants(df, NAME_COL_VARIANTS)
    col_paterno = get_col_by_variants(df, PATERNO_COL_VARIANTS)
    col_materno = get_col_by_variants(df, MATERNO_COL_VARIANTS)

    # Caso 1: Tenemos las 3 columnas separadas (Ideal)
    if col_nombre and col_paterno and col_materno:
        print(f"  ✓ {df_name}: Detectadas columnas separadas: {col_nombre}, {col_paterno}, {col_materno}")
        key_series = (clean_string_for_key(df[col_nombre]) + "|" +
                    clean_string_for_key(df[col_paterno]) + "|" +
                    clean_string_for_key(df[col_materno]))
        return key_series, "Nombre + Paterno + Materno", [col_nombre, col_paterno, col_materno]

    # Caso 2: Nombre y Paterno (Sin Materno)
    if col_nombre and col_paterno:
        print(f"  ⚠ {df_name}: Falta apellido materno. Usando: {col_nombre}, {col_paterno}")
        key_series = (clean_string_for_key(df[col_nombre]) + "|" +
                    clean_string_for_key(df[col_paterno]))
        return key_series, "Nombre + Paterno", [col_nombre, col_paterno]

    # Caso 3: Nombre Completo en una sola columna
    col_full = get_col_by_variants(df, FULLNAME_COL_VARIANTS)
    if col_full:
        print(f"  ✓ {df_name}: Usando columna de nombre completo: '{col_full}'")
        key_series = clean_string_for_key(df[col_full])
        return key_series, "Nombre Completo", [col_full]

    return None, "", []


def extract_duplicate_differences(df_duplicados: pd.DataFrame, key_col: str = '__KEY__') -> pd.DataFrame:
    """
    Extrae solo las diferencias entre registros duplicados.
    Para cada grupo de duplicados, solo muestra las columnas donde hay diferencias.
    Descompone la clave combinada en NOMBRE, APELLIDO_PATERNO, APELLIDO_MATERNO.

    Args:
        df_duplicados: DataFrame con registros duplicados
        key_col: Nombre de la columna clave

    Returns:
        DataFrame con columnas: NOMBRE, APELLIDO_PATERNO, APELLIDO_MATERNO, columnas con diferencias
    """
    if df_duplicados.empty:
        return pd.DataFrame()

    diferencias_list = []

    # Agrupar por la clave para procesar cada grupo de duplicados
    grupos = df_duplicados.groupby(key_col)

    for clave, grupo in grupos:
        if len(grupo) > 1:  # Solo procesar si hay más de un registro
            # Identificar columnas con diferencias
            cols_con_diferencias = []
            for col in grupo.columns:
                if col != key_col and col != '__HOJA_ORIGEN__':
                    # Verificar si hay valores diferentes en esta columna
                    valores_unicos = grupo[col].nunique()
                    if valores_unicos > 1:
                        cols_con_diferencias.append(col)

            # Si hay diferencias, agregar al resultado
            if cols_con_diferencias:
                for idx, row in grupo.iterrows():
                    diff_dict = {}

                    # Descomponer la clave combinada (NOMBRE|PATERNO|MATERNO)
                    if '|' in str(clave):
                        partes = str(clave).split('|')
                        if len(partes) >= 1:
                            diff_dict['NOMBRE'] = partes[0].strip() if partes[0] else ''
                        if len(partes) >= 2:
                            diff_dict['APELLIDO_PATERNO'] = partes[1].strip() if partes[1] else ''
                        if len(partes) >= 3:
                            diff_dict['APELLIDO_MATERNO'] = partes[2].strip() if partes[2] else ''
                    else:
                        # Si no está combinada, asumir que es solo nombre
                        diff_dict['NOMBRE'] = str(clave).strip()

                    diff_dict['_INDICE_'] = idx  # Para identificar cada registro
                    for col in cols_con_diferencias:
                        diff_dict[col] = row[col]
                    diferencias_list.append(diff_dict)

    if diferencias_list:
        return pd.DataFrame(diferencias_list)
    else:
        return pd.DataFrame()


def get_memory_usage(df: pd.DataFrame) -> str:
    """Calcula el uso de memoria de un DataFrame en MB"""
    memory_mb = df.memory_usage(deep=True).sum() / (1024 * 1024)
    return f"{memory_mb:.2f} MB"


def abrir_archivo_xlsx(ruta_archivo: str):
    """Abre un archivo xlsx con la aplicación predeterminada del sistema"""
    try:
        if os.name == 'nt':  # Windows
            os.startfile(ruta_archivo)
        elif os.name == 'posix':  # macOS/Linux
            sistema = platform.system()
            if sistema == 'Darwin':  # macOS
                subprocess.call(['open', ruta_archivo])
            else:  # Linux
                subprocess.call(['xdg-open', ruta_archivo])
        print("\nAbriendo archivo...")
    except OSError as e:
        print(f"\nNo se pudo abrir el archivo automáticamente: {e}")
        print(f"   Por favor, ábrelo manualmente desde: {ruta_archivo}")


def check_system_memory():
    """Verifica la memoria disponible del sistema (Windows) - Requiere psutil (opcional)"""
    try:
        import psutil
        memory = psutil.virtual_memory()
        available_gb = memory.available / (1024 ** 3)
        total_gb = memory.total / (1024 ** 3)
        used_percent = memory.percent
        return {
            'available_gb': available_gb,
            'total_gb': total_gb,
            'used_percent': used_percent
        }
    except (ImportError, ModuleNotFoundError):
        # psutil no está instalado - funcionalidad opcional deshabilitada
        return None
    except (OSError, ValueError, RuntimeError):
        return None


def show_memory_warning(df_size_mb: float):
    """Muestra advertencia si el archivo es muy grande (requiere psutil - opcional)"""
    try:
        memory_info = check_system_memory()

        if memory_info and df_size_mb > 500:
            available_gb = memory_info['available_gb']
            if available_gb < 2:
                print(f"  ⚠️ ADVERTENCIA: Memoria disponible baja ({available_gb:.1f} GB)")
                print("     Se recomienda cerrar otras aplicaciones.")
    except (OSError, ValueError, RuntimeError):
        # Si hay cualquier error, simplemente no mostrar advertencia
        pass


def clear_screen():
    """Limpia la pantalla según el SO - DESHABILITADO POR SOLICITUD"""



def print_header():
    """Imprime encabezado del programa"""
    print("\n" + "="*70)
    print(" "*15 + "🔍 COMPARADOR DE ARCHIVOS XLSX/CSV")
    print("="*70 + "\n")


def seleccionar_archivo_ventana(titulo_ventana: str) -> str:
    """Abre una ventana del sistema para elegir el archivo"""
    root = tk.Tk()
    root.withdraw()  # Oculta la ventana principal pequeña de Tkinter
    root.attributes('-topmost', True)  # Fuerza que la ventana aparezca encima de todo

    ruta_archivo = filedialog.askopenfilename(
        title=titulo_ventana,
        filetypes=[("Archivos Excel y CSV", "*.xlsx *.xls *.csv"), ("Todos los archivos", "*.*")]
    )

    root.destroy()  # Cierra la instancia de tkinter
    return ruta_archivo


def seleccionar_archivos_ventana_multiple(titulo_ventana: str) -> list:
    """Abre una ventana del sistema para elegir múltiples archivos"""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    rutas = filedialog.askopenfilenames(
        title=titulo_ventana,
        filetypes=[("Archivos Excel y CSV", "*.xlsx *.xls *.csv"), ("Todos los archivos", "*.*")]
    )
    root.destroy()
    return list(rutas)


def list_files_in_directory(directory: str = '.', extensions: Optional[list] = None) -> list:
    """Lista archivos en el directorio actual"""
    files = []
    try:
        # Verificar si el directorio existe
        if not os.path.isdir(directory):
            raise ValueError(f"❌ El directorio '{directory}' no existe.")

        for file in os.listdir(directory):
            if os.path.isfile(os.path.join(directory, file)):
                if extensions is None or any(file.lower().endswith(ext) for ext in extensions):
                    files.append(file)
    except OSError as e:
        print(f"❌ Error listando archivos: {e}")
    return sorted(files)


def get_xlsx_sheets(path: str) -> list:
    """Obtiene todas las hojas del archivo XLSX"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except (OSError, ValueError, KeyError, IndexError, TypeError) as e: # Catching expected errors from file I/O or openpyxl
        print(f"❌ Error al leer hojas: {e}")
        return []


def load_all_sheets(path: str) -> dict:
    """Carga todas las hojas de un archivo XLSX y retorna un diccionario {nombre_hoja: DataFrame} (optimizado para archivos grandes)"""
    sheets_dict = {}
    try:
        sheets = get_xlsx_sheets(path)
        if not sheets:
            return sheets_dict

        file_size_mb = os.path.getsize(path) / (1024 * 1024)

        print(f"  📄 Procesando {len(sheets)} hoja(s) ({file_size_mb:.2f} MB)...")

        if file_size_mb > 8 and len(sheets) > 1:
            print("  ⚡ Archivo grande con múltiples hojas - procesamiento optimizado")

        total_rows = 0
        for i, sheet_name in enumerate(sheets, 1):
            print(f"    [{i}/{len(sheets)}] Cargando '{sheet_name}'...", end=' ')
            df = load_table(path, sheet_name=sheet_name)
            if not df.empty:
                sheets_dict[sheet_name] = df
                total_rows += len(df)
                print(f"✓ ({len(df):,} filas)")
            else:
                print("⚠ (vacía)")

        print(f"  ✓ Total de filas cargadas: {total_rows:,}")
        return sheets_dict
    except (OSError, ValueError, KeyError, IndexError, TypeError) as e:
        print(f"❌ Error cargando todas las hojas: {e}")
        return sheets_dict



def select_sheet_interactive(file_path: str) -> Optional[str]:
    """Permite seleccionar una hoja del archivo XLSX o procesar todas"""
    sheets = get_xlsx_sheets(file_path)

    if not sheets:
        return None

    if len(sheets) == 1:
        print(f"  📄 Hoja detectada: '{sheets[0]}'")
        return sheets[0]

    print(f"\n  📄 Hojas disponibles en '{os.path.basename(file_path)}':")
    for idx, sheet in enumerate(sheets, 1):
        print(f"    {idx}. {sheet}")
    print("    4. Usar la primera hoja")
    print("    0. Analizar TODAS las hojas")

    try:
        choice = input("  Seleccione opcion: ").strip()

        if choice.upper() == '0':
            return 'ALL_SHEETS'  # Marcador especial para procesar todas las hojas

        choice_num = int(choice)
        if choice_num == 0 or choice_num == 1:
            return sheets[0]
        elif 1 < choice_num <= len(sheets):
            return sheets[choice_num - 1]
    except ValueError:
        pass

    return sheets[0]


def ask_identification_mode():
    """Configura el modo de identificación para usar RUT, Nombre y Apellidos."""
    print("\n🔍 CONFIGURACIÓN DE CRITERIOS DE BÚSQUEDA")
    print("=" * 70)
    fields = ['rut', 'nombre', 'paterno', 'materno']
    print(f"  ✓ Modo de identificación fijado a: {', '.join(fields).upper()}")
    
    config = {
        'mode': 'manual',
        'fields': fields
    }
    
    input("\nPresione Enter para continuar...")
    return config


def interactive_menu_individual_selection() -> Tuple[str, str, Optional[str], Optional[str], Optional[str], Optional[list], dict]:
    """Menú interactivo seleccionando archivos uno por uno"""
    clear_screen()
    print_header()

    # CONFIGURACIÓN PREVIA (Solicitud)
    config_ident = ask_identification_mode()
    if config_ident is None:
        return "", "", None, None, None, [], {}

    clear_screen()
    print_header()

    print("📋 PASO 1: Seleccionar Archivos")
    print("=" * 70)

    print("\n1️⃣ Abriendo ventana para seleccionar el PRIMER archivo (Examinado)...")
    path_a = seleccionar_archivo_ventana("Selecciona el archivo EXAMINADO (Base)")
    if not path_a:
        return "", "", None, None, None, [], {}
    print(f"  ✓ Archivo A: {os.path.basename(path_a)}")

    print("\n2️⃣ Abriendo ventana para seleccionar el SEGUNDO archivo (Ejecución)...")
    path_b = seleccionar_archivo_ventana("Selecciona el archivo EJECUCIÓN (Comparar)")
    if not path_b:
        return "", "", None, None, None, [], {}
    print(f"  ✓ Archivo B: {os.path.basename(path_b)}")

    # Por defecto seleccionamos todos los análisis, luego se filtra al guardar
    selected_analysis_types = None

    # Seleccionar hoja A
    selected_sheet_a = None
    if path_a.lower().endswith(('.xlsx', '.xls')):
        selected_sheet_a = select_sheet_interactive(path_a)

    # Seleccionar hoja B
    selected_sheet_b = None
    if path_b.lower().endswith(('.xlsx', '.xls')):
        selected_sheet_b = select_sheet_interactive(path_b)

    # Columna clave (detección automática)
    clear_screen()
    print_header()
    print("📋 PASO 2: Configurar Columna Clave")
    print("=" * 70)

    print("\n✓ Configuración lista.")

    selected_key = None

    return path_a, path_b, selected_key, selected_sheet_a, selected_sheet_b, selected_analysis_types, config_ident


def interactive_menu() -> Tuple[str, str, Optional[str], Optional[str], Optional[str], Optional[list], dict]:
    """Menú interactivo para seleccionar archivos y parámetros con ventanas"""
    clear_screen()
    print_header()

    # CONFIGURACIÓN PREVIA (Solicitud)
    config_ident = ask_identification_mode()
    if config_ident is None:
        return "", "", None, None, None, [], {}

    clear_screen()
    print_header()

    print("📋 PASO 1: Seleccionar Archivos")
    print("=" * 70)

    print("\n1️⃣ Abriendo ventana para seleccionar Archivos (puedes elegir varios a la vez)...")
    archivos = seleccionar_archivos_ventana_multiple("Selecciona las Bases de Datos")

    if len(archivos) < 2:
        print("❌ Debes seleccionar al menos 2 archivos.")
        return "", "", None, None, None, [], {}

    path_a, path_b = archivos[0], archivos[1]
    print(f"  ✓ Archivo A: {os.path.basename(path_a)}")
    print(f"  ✓ Archivo B: {os.path.basename(path_b)}")

    # Por defecto seleccionamos todos los análisis, luego se filtra al guardar
    selected_analysis_types = None

    # Seleccionar hoja A
    selected_sheet_a = None
    if path_a.lower().endswith(('.xlsx', '.xls')):
        selected_sheet_a = select_sheet_interactive(path_a)

    # Seleccionar hoja B
    selected_sheet_b = None
    if path_b.lower().endswith(('.xlsx', '.xls')):
        selected_sheet_b = select_sheet_interactive(path_b)

    # Columna clave (detección automática)
    clear_screen()
    print_header()
    print("📋 PASO 2: Configurar Columna Clave")
    print("=" * 70)

    print("\n✓ Configuración lista.")

    selected_key = None

    return path_a, path_b, selected_key, selected_sheet_a, selected_sheet_b, selected_analysis_types, config_ident


def detect_header_row_xlsx(path: str, sheet_name: Optional[str] = None) -> Tuple[int, list]:
    """
    Detecta automáticamente la fila de encabezado en XLSX
    Retorna: (índice_fila, lista_columnas)
    Optimizado para archivos grandes
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Validación: asegúrate de que ws no es None
        if ws is None:
            print(f"⚠ No se pudo acceder a la hoja en {path}")
            wb.close()
            return 0, []

        # Validación: verificar si la hoja está vacía
        if ws.max_row is None or ws.max_row == 0:
            print(f"⚠ La hoja '{sheet_name or 'activa'}' está vacía")
            wb.close()
            return 0, []

        key_set = {k.lower() for k in COMMON_KEY_NAMES}
        # Limitar verificación para no afectar performance en archivos grandes
        max_check_rows = min(20, ws.max_row)

        best_match = (-1, 0, [])

        for i in range(1, max_check_rows + 1):
            # CORRECCIÓN: Validar que iter_rows devuelva datos
            rows_iter = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))

            # Validar que hay al menos una fila
            if not rows_iter or len(rows_iter) == 0:
                continue

            row = rows_iter[0]

            # Validar que la fila no esté vacía
            if not row or all(v is None or str(v).strip() == '' for v in row):
                continue

            row_vals = [str(v).strip().lower() if v else '' for v in row]

            # Contar coincidencias con claves comunes
            matches = sum(1 for v in row_vals if v in key_set)

            # Verificar si hay datos válidos (no vacíos)
            non_empty = sum(1 for v in row_vals if v and v.strip())

            if matches > best_match[1] and non_empty > 0:
                best_match = (i - 1, matches, row_vals)

        # Estrategia 1: Si encontramos encabezados por coincidencia clara, usarlos
        if best_match[0] >= 0 and best_match[1] > 0:
            wb.close()
            return best_match[0], best_match[2]

        # ESTRATEGIA SECUNDARIA: Detección por densidad
        # Si no encontramos palabras clave, buscamos la primera fila con más columnas
        print("  ℹ No se detectaron palabras clave en encabezados. Buscando por densidad de datos...")
        max_cols = 0
        best_density_row = 0

        # IMPORTANTE: No cerrar el workbook todavía porque lo usamos aquí
        for i in range(1, min(15, ws.max_row)):
            row_vals = list(ws.iter_rows(min_row=i, max_row=i, values_only=True))[0]
            non_empty = sum(1 for v in row_vals if v is not None and str(v).strip() != '')

            # Si encontramos una fila con significativamente más columnas
            if non_empty > max_cols:
                max_cols = non_empty
                best_density_row = i - 1

        # Si la fila detectada por densidad tiene al menos 2 columnas, usarla
        if max_cols >= 2:
            print(f"  ✓ Encabezado detectado por estructura en fila {best_density_row + 1}")
            # Obtener nombres de columnas
            cols = list(ws.iter_rows(min_row=best_density_row+1, max_row=best_density_row+1, values_only=True))[0]
            wb.close()
            return best_density_row, list(cols)

        wb.close()
        # Si no se encontró encabezado, retornar 0
        print(f"⚠ No se detectó encabezado claro en '{sheet_name or 'hoja activa'}', usando fila 1")
        return 0, []

    except (OSError, ValueError, KeyError, IndexError, TypeError) as e:
        print(f"❌ Error detectando encabezado: {e}")
        return 0, []


def load_table(path: str, sheet_name: Optional[str] = None, auto_detect: bool = True, use_chunks: bool = False) -> pd.DataFrame:
    """Carga tabla de XLSX/CSV con detección automática de encabezado y optimización para archivos grandes (>8MB)"""
    ext = os.path.splitext(path)[1].lower()

    # Verificar tamaño del archivo
    file_size_mb = os.path.getsize(path) / (1024 * 1024)
    is_large_file = file_size_mb > 8

    if is_large_file:
        print(f"  📦 Archivo grande detectado: {file_size_mb:.2f} MB - aplicando optimizaciones...")

    try:
        if ext in ['.xls', '.xlsx']:
            # Detectar encabezado automáticamente
            header_idx, _ = detect_header_row_xlsx(path, sheet_name)

            # Configurar parámetros de lectura
            read_params = {
                'io': path,
                'header': header_idx if (auto_detect and header_idx > 0) else 0,
                'dtype': str,
                'engine': 'openpyxl'
            }

            # Agregar sheet_name solo si no es None
            if sheet_name is not None:
                read_params['sheet_name'] = sheet_name

            df = pd.read_excel(**read_params)

            # Verificar que df es un DataFrame (no un diccionario)
            if isinstance(df, dict):
                df = next(iter(df.values()))

            # CORRECCIÓN: Convertir columnas a string antes de usar .str
            # Manejar columnas sin nombre (None o NaN)
            df.columns = [str(col).strip() if col is not None and str(col) != 'nan'
                         else f'Columna_{i}' for i, col in enumerate(df.columns)]

            # IMPORTANTE: fillna ANTES de convertir a category
            df = df.fillna('')

            # Optimización agresiva para archivos grandes
            if is_large_file:
                print("  🔧 Optimizando tipos de datos para reducir memoria...")
                for col in df.columns:
                    if len(df) > 0:
                        unique_ratio = df[col].nunique() / len(df)
                        # Usar categorías para columnas con menos del 40% de valores únicos
                        if unique_ratio < 0.4:
                            try:
                                df[col] = df[col].astype('category')
                            except (ValueError, TypeError):
                                pass
            elif len(df) > 10000:
                # Optimización estándar para archivos medianos
                for col in df.columns:
                    unique_ratio = df[col].nunique() / len(df)
                    if unique_ratio < 0.5:
                        try:
                            df[col] = df[col].astype('category')
                        except (ValueError, TypeError):
                            pass

            return df

        if ext == '.csv':
            # Para CSV grandes (>8MB), siempre usar chunks
            if is_large_file or use_chunks:
                print("  📊 Procesando archivo CSV por bloques (chunks)...")
                chunks = []
                chunk_size = 30000 if is_large_file else 50000
                for i, chunk in enumerate(pd.read_csv(path, dtype=str, keep_default_na=False,
                                        engine='c', chunksize=chunk_size), 1):
                    chunks.append(chunk)
                    if i % 5 == 0:
                        print(f"    Procesados {i * chunk_size:,} registros...")
                df = pd.concat(chunks, ignore_index=True)
                print(f"  ✓ Total cargado: {len(df):,} registros")
            else:
                df = pd.read_csv(path, dtype=str, keep_default_na=False, engine='c', low_memory=False)

            # CORRECCIÓN: Igual para CSV
            df.columns = [str(col).strip() if col is not None and str(col) != 'nan'
                         else f'Columna_{i}' for i, col in enumerate(df.columns)]
            df = df.fillna('')
            return df

        raise ValueError(f"❌ Formato no soportado: {ext}")

    except (OSError, ValueError, KeyError, IndexError, TypeError) as e:
        print(f"❌ Error cargando {path}: {e}")
        return pd.DataFrame()


def auto_detect_key_column(df: pd.DataFrame, provided_key: Optional[str] = None) -> str:
    """
    Detecta automáticamente la columna clave.
    Prioridad: parámetro > nombres comunes > primera columna numérica > primera columna
    """
    if provided_key and provided_key in df.columns:
        return provided_key

    # Buscar nombres comunes
    for name in COMMON_KEY_NAMES:
        for col in df.columns:
            if col.lower() == name.lower():
                return col

    # Buscar columna que parezca identificador (menos valores nulos)
    non_null_counts = (df != '').sum()
    best_col = non_null_counts.idxmax()

    # Asegurar que devolvemos un string
    return str(best_col)


def analyze_column_uniqueness(df: pd.DataFrame, col: str) -> dict:
    """Analiza la unicidad de una columna (optimizado para grandes volúmenes)"""
    # Usar operaciones vectorizadas para mejor performance
    values = df[col].astype(str).str.strip()
    total = len(df)
    unique = values.nunique()
    null_count = (values == '').sum()
    duplicates = total - unique

    return {
        'total': total,
        'unique': unique,
        'duplicates': duplicates,
        'null': null_count,
        'uniqueness_pct': (unique / total * 100) if total > 0 else 0
    }


def find_matching_key_columns(df_a: pd.DataFrame, df_b: pd.DataFrame,
                              key_a: str, key_b: str) -> Tuple[str, str]:
    """
    Encuentra las mejores columnas clave para comparación.
    Si las columnas originales no coinciden bien, busca alternativas.
    """
    stats_a = analyze_column_uniqueness(df_a, key_a)
    stats_b = analyze_column_uniqueness(df_b, key_b)

    # Si ambas columnas tienen buena unicidad, usarlas
    if stats_a['uniqueness_pct'] > 90 and stats_b['uniqueness_pct'] > 90:
        return key_a, key_b

    # Buscar mejor coincidencia
    best_score = -1
    best_pair = (key_a, key_b)

    for col_a in df_a.columns[:10]:  # Limitar búsqueda a primeras 10 columnas
        stats_a_alt = analyze_column_uniqueness(df_a, col_a)

        for col_b in df_b.columns[:10]:
            stats_b_alt = analyze_column_uniqueness(df_b, col_b)

            # Score: suma de unicidades
            score = stats_a_alt['uniqueness_pct'] + stats_b_alt['uniqueness_pct']

            if score > best_score:
                best_score = score
                best_pair = (col_a, col_b)

    return best_pair


def mark_incomplete(df: pd.DataFrame, exclude_cols: Optional[list] = None) -> pd.DataFrame:
    """Marca filas con datos incompletos"""
    if exclude_cols is None:
        exclude_cols = []

    cols_check = [c for c in df.columns if c not in exclude_cols]
    mask = (df[cols_check] == '').any(axis=1)
    return df[mask]


def format_rut(rut_str: str) -> str:
    """
    Formatea un RUT chileno al formato XX.XXX.XXX-X
    Ej: 163456789 -> 16.345.678-9
    Ej: 15811.479-8 -> 15.811.479-8
    """
    if not rut_str or rut_str == '':
        return ''

    rut_str = str(rut_str).strip()

    # Extraer solo dígitos y K (dígito verificador puede ser K)
    rut_limpio = ''
    for c in rut_str:
        if c.isdigit():
            rut_limpio += c
        elif c.upper() == 'K':
            rut_limpio += c.upper()

    if len(rut_limpio) < 2:
        return rut_str

    # Separar cuerpo y dígito verificador
    # El dígito verificador está al final (puede ser número o K)
    digito = rut_limpio[-1]
    cuerpo = rut_limpio[:-1]

    # Formatear el cuerpo con puntos cada 3 dígitos de derecha a izquierda
    cuerpo_formateado = ''
    for i, digit in enumerate(reversed(cuerpo)):
        if i > 0 and i % 3 == 0:
            cuerpo_formateado = '.' + cuerpo_formateado
        cuerpo_formateado = digit + cuerpo_formateado

    return f"{cuerpo_formateado}-{digito}"


def format_dataframe_rut(df: pd.DataFrame, rut_column: str) -> pd.DataFrame:
    """
    Formatea todos los RUTs de una columna específica
    """
    df_copy = df.copy()
    if rut_column in df_copy.columns:
        df_copy[rut_column] = df_copy[rut_column].apply(format_rut)
    return df_copy


def find_null_data_columns(df: pd.DataFrame, exclude_cols: Optional[list] = None) -> dict:
    """
    Identifica todas las columnas que contienen datos nulos/vacíos
    Retorna un diccionario con información detallada de nulidades
    """
    if exclude_cols is None:
        exclude_cols = []

    null_info = {
        'columnas_con_nulos': [],
        'cantidad_nulos_por_columna': {},
        'porcentaje_nulos_por_columna': {},
        'total_celdas_nulas': 0
    }

    for col in df.columns:
        if col not in exclude_cols and col != '__KEY__':
            # Contar nulos (valores vacíos)
            null_count = (df[col] == '').sum()
            if null_count > 0:
                null_info['columnas_con_nulos'].append(col)
                null_info['cantidad_nulos_por_columna'][col] = int(null_count)
                null_info['porcentaje_nulos_por_columna'][col] = (null_count / len(df) * 100)
                null_info['total_celdas_nulas'] += int(null_count)

    return null_info


def print_null_stats_table(null_info: dict, title: str):
    """Imprime una tabla formateada con estadísticas de nulos"""
    print(f"\n📊 {title}")
    print(f"  Columnas con datos nulos: {len(null_info['columnas_con_nulos'])}")
    print(f"  Total celdas nulas: {null_info['total_celdas_nulas']:,}")

    if null_info['columnas_con_nulos']:
        # Encabezado de la tabla
        print(f"\n  {'COLUMNA':<50} | {'CANTIDAD':>10} | {'% NULOS':>8}")
        print("  " + "-"*76)

        # Preparar datos para ordenar
        data = []
        for col in null_info['columnas_con_nulos']:
            qty = null_info['cantidad_nulos_por_columna'][col]
            pct = null_info['porcentaje_nulos_por_columna'][col]
            data.append((col, qty, pct))

        # Ordenar por porcentaje descendente (los más críticos primero)
        data.sort(key=lambda x: x[2], reverse=True)

        for col, qty, pct in data:
            # Truncar nombre si es muy largo
            col_display = (col[:47] + '...') if len(col) > 47 else col
            print(f"  {col_display:<50} | {qty:>10,} | {pct:>7.2f}%")
        print("  " + "-"*76)


def crear_reporte_datos_faltantes(df: pd.DataFrame, key_column: str, _output_dir: str = '.'):
    """
    Crea un reporte detallado de qué datos están nulos por usuario (identificado por RUT/key)
    Formatea automáticamente los RUTs al formato chileno
    """
    # Filtrar solo filas con al menos un dato nulo
    mask = (df != '').sum(axis=1) < len(df.columns) - 1  # Al menos una columna vacía
    df_with_nulls = df[mask].copy()

    if df_with_nulls.empty:
        return None

    # Crear reporte detallado
    reporte = []

    for _, row in df_with_nulls.iterrows():
        usuario_key = row[key_column]
        campos_nulos = []

        for col in df.columns:
            if col != '__KEY__' and col != key_column:
                if row[col] == '':
                    campos_nulos.append(col)

        if campos_nulos:
            reporte.append({
                key_column: format_rut(usuario_key),  # Formatear RUT
                'Campos_Nulos': ', '.join(campos_nulos),
                'Cantidad_Campos_Faltantes': len(campos_nulos)
            })

    if not reporte:
        return None

    df_reporte = pd.DataFrame(reporte)
    df_reporte = df_reporte.sort_values('Cantidad_Campos_Faltantes', ascending=False)

    return df_reporte


def print_progress(label: str, percent: float, width: int = 20):
    """Imprime una barra de progreso en consola."""
    filled = int(width * percent / 100)
    bar = '█' * filled + '░' * (width - filled)
    print(f"  {label}: [{bar}] {percent:5.1f}%", end='\r', flush=True)


def search_in_dataframe(df: pd.DataFrame, term: str) -> pd.DataFrame:
    """Busca un término en un DataFrame insensible a mayúsculas"""
    try:
        mask = pd.DataFrame(False, index=df.index, columns=df.columns)
        cols_obj = df.select_dtypes(include=['object', 'string']).columns

        for col in cols_obj:
            mask[col] = df[col].astype(str).str.upper().str.contains(term, na=False, regex=False)

        return df[mask.any(axis=1)]
    except (ValueError, TypeError, AttributeError):
        return pd.DataFrame()


def print_dataframe_table(df_input: pd.DataFrame, max_rows: int = 10):
    """Imprime un DataFrame usando tabulate si es posible"""
    try:
        # Mostrar un resumen (primeras columnas para identificar)
        cols_prio = [c for c in df_input.columns if any(x in c.lower() for x in ['nombre', 'paterno', 'materno', 'rut', 'id', 'centro', 'comuna'])]
        cols_rest = [c for c in df_input.columns if c not in cols_prio]
        cols_show = (cols_prio + cols_rest)[:8]

        from tabulate import tabulate
        # Preparar datos para tabulate
        df_print = df_input[cols_show].copy()
        df_print = df_print.fillna('')
        df_print = df_print.astype(str)

        # Truncar columnas largas
        for col in df_print.columns:
            df_print[col] = df_print[col].str.slice(0, 25)

        # Ajustar nombres de columnas
        df_print.columns = [str(c)[:15] for c in df_print.columns]

        headers = df_print.columns.tolist()

        if len(df_input) > max_rows:
            data = df_print.head(max_rows).values.tolist()
            print(tabulate(data, headers=headers, tablefmt='grid', showindex=False))
            print(f"      ... y {len(df_input)-max_rows} filas más.")
        else:
            data = df_print.values.tolist()
            print(tabulate(data, headers=headers, tablefmt='grid', showindex=False))

    except ImportError:
        print("-" * 100)
        print(df_input.to_string(index=False))
        print("-" * 100)
    except (ValueError, TypeError, AttributeError) as e:
        print(f"Error formato: {e}")
        print("-" * 40)
        print(df_input.head(max_rows).to_string(index=False))
        if len(df_input) > max_rows:
            print(f"      ... y {len(df_input)-max_rows} filas más.")
        print("-" * 40)


def save_outputs_single_file(reportes_dict: dict, output_dir: str = '.', selected_analysis_types: Optional[list] = None):
    """
    Guarda reportes en un archivo Excel según los tipos de análisis seleccionados

    Args:
        reportes_dict: Diccionario con los DataFrames de reportes
        output_dir: Directorio donde guardar el archivo
        selected_analysis_types: Lista con tipos ['duplicados', 'faltantes', 'incompletos'] o None para todos
    """


    # -------------------------------------------------------------
    # DIAGNÓSTICO E INTERACCIÓN PREVIA
    # -------------------------------------------------------------
    print("\n" + "="*60)
    print("📋  DIAGNÓSTICO DE RESULTADOS ENCONTRADOS")
    print("="*60)

    has_results = False

    # Obtener totales para cálculos de porcentaje
    total_a = reportes_dict.get('_TOTAL_A', 0)
    total_b = reportes_dict.get('_TOTAL_B', 0)
    nombre_a = reportes_dict.get('_NOMBRE_A', '')
    nombre_b = reportes_dict.get('_NOMBRE_B', '')

    # Definir orden de prioridad para los reportes
    priority_order = [
        '_COMPLETO_CON_MARCADORES',
        'TODOS - Faltantes',
        'Faltantes en ',
        'TODOS - Duplicados',
        'Duplicados en ',
        'TODOS - Incompletos',
        'Incompletos en ',
        'DIAGNOSTICO_PRIORITARIO'
    ]
    
    def get_priority(name):
        """Devuelve la prioridad de un reporte basado en su nombre"""
        for idx, pattern in enumerate(priority_order):
            if pattern in name:
                return idx
        return len(priority_order)  # Reportes sin prioridad van al final
    
    # Ordenar reportes por prioridad
    sorted_reports = sorted(
        [(k, v) for k, v in reportes_dict.items() if not k.startswith('_') and isinstance(v, pd.DataFrame)],
        key=lambda x: (get_priority(x[0]), x[0])
    )
    
    for report_name, val in sorted_reports:
        # Omitir claves genéricas "en A" y "en B" si tienen nombres específicos
        if (report_name.endswith(" en A") and nombre_a != "A") or \
        (report_name.endswith(" en B") and nombre_b != "B"):
            continue

        count = len(val)
        icon = "✅" if count > 0 else "⚪"

        # Calcular porcentaje si aplica
        pct_str = ""
        if nombre_a and (f"en {nombre_a}" in report_name or "en A" in report_name):
            if total_a > 0:
                pct = (count / total_a) * 100
                pct_str = f" ({pct:5.2f}% de {nombre_a})"
        elif nombre_b and (f"en {nombre_b}" in report_name or "en B" in report_name):
            if total_b > 0:
                pct = (count / total_b) * 100
                pct_str = f" ({pct:5.2f}% de {nombre_b})"

        print(f"  {icon} {report_name:<30}: {count:>6} registros{pct_str}")
        if count > 0:
            has_results = True

    print("-" * 60)

    if not has_results:
        print("\n⚠  ATENCIÓN: No se encontraron diferencias ni datos para reportar.")

    # === FUNCIONALIDAD DE BÚSQUEDA DE USUARIOS (Solicitada) ===
    # Flexible: por nombre, apellidos o RUT
    while True:
        print("\n" + "-"*60)
        print("🔍 BÚSQUEDA INTERACTIVA DE USUARIOS")
        print("-" * 60)
        print("Puedes buscar por: Nombre, Apellido Paterno, Materno o RUT")
        resp_buscar = input("¿Deseas buscar algún usuario en los resultados? (s/n): ").strip().lower()

        if resp_buscar == 'n' or resp_buscar == 'no':
            break

        if resp_buscar == 's' or resp_buscar == 'si' or resp_buscar == 'y':
            termino = input("\n   ✍️  Ingrese término de búsqueda: ").strip().upper()
            if not termino:
                continue

            print(f"\n   ⏳ Buscando '{termino}' en todas las tablas generadas...")
            encontrados_total = 0

            for report_name, df_res in reportes_dict.items():
                if not (isinstance(df_res, pd.DataFrame) and not df_res.empty and not report_name.startswith('_')):
                    continue

                # Usar función auxiliar para búsqueda (reduce nesting y try-except genérico)
                filas_encontradas = search_in_dataframe(df_res, termino)

                if filas_encontradas.empty:
                    continue

                count = len(filas_encontradas)
                encontrados_total += count
                print(f"\n   👉 Encontrado en tabla '{report_name}': {count} coincidencias")

                # Imprimir tabla bien formateada
                print_dataframe_table(filas_encontradas)

            if encontrados_total == 0:
                print(f"\n   ❌ No se encontraron registros que contengan '{termino}'.")
            else:
                print(f"\n   ✅ Total coincidencias encontradas: {encontrados_total}")

            input("\n   Presione Enter para continuar buscando o 'n' en la próxima pregunta...")
        else:
            print("   ⚠️ Opción no válida. Ingrese 's' para buscar o 'n' para continuar.")

    print("\n¿Qué datos deseas descargar?")
    print("  1. Duplicados")
    print("  2. Incompletos")
    print("  3. Faltantes")
    print("  4. Todos (Duplicados + Incompletos + Faltantes)")
    print("  0. Volver al menú principal")
    print("  x. Detener programa")
    print("="*60)

    while True:
        seleccion = input("\nEscribe tu opción (1-4, 0, x): ").strip().lower()

        if seleccion == '0':
            print("\n🔙 Volviendo al menú principal...")
            return None

        if seleccion == 'x':
            print("\n👋 Deteniendo programa...")
            sys.exit()

        if seleccion == '1':
            selected_analysis_types = ['duplicados']
            break
        elif seleccion == '2':
            selected_analysis_types = ['incompletos']
            break
        elif seleccion == '3':
            selected_analysis_types = ['faltantes']
            break
        elif seleccion == '4':
            selected_analysis_types = ['duplicados', 'incompletos', 'faltantes']
            break
        else:
            print("❌ Opción no válida. Intenta de nuevo.")

    # Determinar nombre del archivo según la selección
    if len(selected_analysis_types) == 1:
        tipo_nombre = selected_analysis_types[0].upper()
        xlsx_path = os.path.join(output_dir, f'REPORTE_{tipo_nombre}.xlsx')
    else:
        xlsx_path = os.path.join(output_dir, 'REPORTE_COMPLETO_COMPARACION.xlsx')

    print(f"\n  📂 Archivo destino: {os.path.basename(xlsx_path)}")
    print("\n🚀 Iniciando generación de archivo Excel...")

    # Calcular tamaño estimado de datos
    total_rows = sum(len(df) for key, df in reportes_dict.items()
                    if isinstance(df, pd.DataFrame) and not key.startswith('_'))

    if total_rows > 50000:
        print(f"  📦 Generando reporte grande ({total_rows:,} filas totales)...")
        print("  ⏳ Esto puede tomar unos minutos...")
        print_progress("Progreso", 0)

    try:
        # Importar estilos
        from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

        # Definir bordes y estilos
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        null_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        null_font = Font(color="FFFFFF")

        # Crear workbook
        wb = openpyxl.Workbook()
        if wb.active is not None:
            wb.remove(wb.active)  # Eliminar hoja por defecto

        # Recuperar nombres de archivos para títulos dinámicos
        nombre_a = reportes_dict.get('_NOMBRE_A', 'A')
        nombre_b = reportes_dict.get('_NOMBRE_B', 'B')

        # Función auxiliar para escribir DataFrame y aplicar formato
        def write_and_format_dataframe(ws, df, start_row, start_col, titulo=None):
            """Escribe un DataFrame en la hoja y aplica formato, retorna la columna final"""

            # Si el DataFrame está vacío, escribir solo el título
            if df.empty:
                if titulo:
                    title_cell = ws.cell(row=start_row, column=start_col, value=titulo + " (Sin datos)")
                    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    title_cell.alignment = Alignment(horizontal='center', vertical='center')
                    title_cell.border = thin_border
                return start_col + 1

            # Formatear RUTs en el DataFrame si existe la columna
            df_display = df.copy()
            rut_columns = [c for c in df_display.columns if c.lower() in ['rut', 'id_rut', 'documento', 'cedula', 'doc']]
            if rut_columns:
                col_rut = rut_columns[0]
                df_display[col_rut] = df_display[col_rut].apply(format_rut)

            # Escribir título si se proporciona
            current_row = start_row
            if titulo:
                # Fusionar celdas para el título
                end_col = start_col + len(df_display.columns) - 1
                try:
                    ws.merge_cells(start_row=current_row, start_column=start_col,
                        end_row=current_row, end_column=end_col)
                except (ValueError, TypeError, AttributeError):
                    pass
                title_cell = ws.cell(row=current_row, column=start_col, value=titulo)
                title_cell.font = Font(bold=True, size=14, color="FFFFFF")
                title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                title_cell.border = thin_border
                current_row += 1

            # Escribir encabezados
            for col_idx, col_name in enumerate(df_display.columns, start=start_col):
                cell = ws.cell(row=current_row, column=col_idx, value=col_name)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # Escribir datos
            for row_idx, row_data in enumerate(df_display.values, start=current_row + 1):
                for col_idx, value in enumerate(row_data, start=start_col):
                    is_null = value is None or str(value).strip() == ''
                    display_value = '-' if is_null else value
                    cell = ws.cell(row=row_idx, column=col_idx, value=display_value)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                    # Aplicar formato rojo para celdas nulas/vacías
                    if is_null:
                        cell.fill = null_fill
                        cell.font = null_font

            # Ajustar ancho de columnas
            for col_idx in range(start_col, start_col + len(df_display.columns)):
                max_length = 0
                column_letter = get_column_letter(col_idx)

                # Calcular desde el título si existe, si no desde los encabezados
                check_start_row = start_row
                for row_idx in range(check_start_row, check_start_row + len(df_display) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (AttributeError, TypeError, ValueError):
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Retorna el próximo col_idx después de esta tabla
            return start_col + len(df_display.columns)

        # Crear hojas con tablas lado a lado
        print("  💾 Generando archivo Excel...")

        # Determinar qué hojas crear según selected_analysis_types
        if selected_analysis_types is None:
            tipos_a_procesar = ['duplicados', 'faltantes', 'incompletos']
        else:
            tipos_a_procesar = selected_analysis_types

        # Contar hojas a crear
        hojas_a_crear = []
        if 'faltantes' in tipos_a_procesar:
            hojas_a_crear.extend(['faltantes', 'todos_faltantes'])
        if 'duplicados' in tipos_a_procesar:
            hojas_a_crear.extend(['duplicados', 'todos_duplicados'])
        if 'incompletos' in tipos_a_procesar:
            hojas_a_crear.extend(['incompletos', 'todos_incompletos'])

        total_hojas = len(hojas_a_crear) + 2  # +2 para usuarios faltantes
        hoja_actual = 0

        # DEBUG: Mostrar qué hay en reportes_dict
        print(f"\n  🔍 Debug - Análisis seleccionados: {tipos_a_procesar}")
        print("  🔍 Debug - Claves en reportes_dict:")
        for report_name in reportes_dict.keys():
            if not report_name.startswith('_'):
                if isinstance(reportes_dict[report_name], pd.DataFrame):
                    print(f"      - {report_name}: {len(reportes_dict[report_name])} filas")
                else:
                    print(f"      - {report_name}: {type(reportes_dict[report_name])}")

        # [MODIFICADO] Se omite la pestaña de Diagnóstico en el Excel (solo se ve en terminal)

        # 1. FALTANTES (solo si está en tipos_a_procesar)
        if 'faltantes' in tipos_a_procesar:
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Primera hoja: TODOS los faltantes consolidados
            if 'TODOS - Faltantes' in reportes_dict and not reportes_dict['TODOS - Faltantes'].empty:
                ws_faltantes = wb.create_sheet("RESUMEN DIFERENCIAS")
                write_and_format_dataframe(ws_faltantes, reportes_dict['TODOS - Faltantes'], 1, 1, "Todos los Registros con Diferencias")
                print(f"    ✓ Creada hoja: RESUMEN DIFERENCIAS ({len(reportes_dict['TODOS - Faltantes'])} filas)")
            hoja_actual += 1
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Hojas adicionales: Faltantes separados por hoja de origen
            for report_name, df in reportes_dict.items():
                if report_name.startswith('Faltantes en ') and report_name not in [f'Faltantes en {nombre_a}', f'Faltantes en {nombre_b}', 'Faltantes en A', 'Faltantes en B']:
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        ws = wb.create_sheet(report_name)
                        write_and_format_dataframe(ws, df, 1, 1, report_name)
                        print(f"    ✓ Creada hoja: {report_name} ({len(df)} filas)")

            hoja_actual += 1
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

        # 2. DUPLICADOS (solo si está en tipos_a_procesar)
        if 'duplicados' in tipos_a_procesar:
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Primera hoja: TODOS los duplicados consolidados
            if 'TODOS - Duplicados' in reportes_dict and not reportes_dict['TODOS - Duplicados'].empty:
                ws_duplicados = wb.create_sheet("TODOS - Duplicados")
                write_and_format_dataframe(ws_duplicados, reportes_dict['TODOS - Duplicados'], 1, 1, "Todos los Registros Duplicados")
                print(f"    ✓ Creada hoja: TODOS - Duplicados ({len(reportes_dict['TODOS - Duplicados'])} filas)")
            hoja_actual += 1
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Segunda hoja: Duplicados en A
            if 'Duplicados en A' in reportes_dict and not reportes_dict['Duplicados en A'].empty:
                ws_dup_a = wb.create_sheet(f"Duplicados en {nombre_a}")
                write_and_format_dataframe(ws_dup_a, reportes_dict['Duplicados en A'], 1, 1, f"Duplicados en {nombre_a}")
                print(f"    ✓ Creada hoja: Duplicados en {nombre_a} ({len(reportes_dict['Duplicados en A'])} filas)")

            # Tercera hoja: Duplicados en B
            if 'Duplicados en B' in reportes_dict and not reportes_dict['Duplicados en B'].empty:
                ws_dup_b = wb.create_sheet(f"Duplicados en {nombre_b}")
                write_and_format_dataframe(ws_dup_b, reportes_dict['Duplicados en B'], 1, 1, f"Duplicados en {nombre_b}")
                print(f"    ✓ Creada hoja: Duplicados en {nombre_b} ({len(reportes_dict['Duplicados en B'])} filas)")

            hoja_actual += 1
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

        # 3. INCOMPLETOS (solo si está en tipos_a_procesar)
        if 'incompletos' in tipos_a_procesar:
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Primera hoja: TODOS los incompletos consolidados
            if 'TODOS - Incompletos' in reportes_dict and not reportes_dict['TODOS - Incompletos'].empty:
                ws_incompletos = wb.create_sheet("TODOS - Incompletos")
                write_and_format_dataframe(ws_incompletos, reportes_dict['TODOS - Incompletos'], 1, 1, "Todos los Registros Incompletos")
                print(f"    ✓ Creada hoja: TODOS - Incompletos ({len(reportes_dict['TODOS - Incompletos'])} filas)")
            hoja_actual += 1
            print_progress("Guardando", (hoja_actual / total_hojas) * 100)

            # Segunda hoja: Incompletos en A
            if 'Incompletos en A' in reportes_dict and not reportes_dict['Incompletos en A'].empty:
                ws_inc_a = wb.create_sheet(f"Incompletos en {nombre_a}")
                write_and_format_dataframe(ws_inc_a, reportes_dict['Incompletos en A'], 1, 1, f"Incompletos en {nombre_a}")
                print(f"    ✓ Creada hoja: Incompletos en {nombre_a} ({len(reportes_dict['Incompletos en A'])} filas)")

            # Tercera hoja: Incompletos en B
            if 'Incompletos en B' in reportes_dict and not reportes_dict['Incompletos en B'].empty:
                ws_inc_b = wb.create_sheet(f"Incompletos en {nombre_b}")
                write_and_format_dataframe(ws_inc_b, reportes_dict['Incompletos en B'], 1, 1, f"Incompletos en {nombre_b}")
                print(f"    ✓ Creada hoja: Incompletos en {nombre_b} ({len(reportes_dict['Incompletos en B'])} filas)")
        print_progress("Guardando", (hoja_actual / total_hojas) * 100)

        if 'Usuarios Faltantes B' in reportes_dict and not reportes_dict['Usuarios Faltantes B'].empty:
            ws = wb.create_sheet("Usuarios Faltantes B")
            write_and_format_dataframe(ws, reportes_dict['Usuarios Faltantes B'], 1, 1, f"Usuarios con Datos Faltantes en {nombre_b}")
        hoja_actual += 1
        print_progress("Guardando", (hoja_actual / total_hojas) * 100)

        # Verificar que se hayan creado hojas
        if len(wb.worksheets) == 0:
            print("\n❌ ERROR: No hay datos para generar el reporte.")
            print("\n   Posibles causas:")
            print("   • El análisis seleccionado no encontró registros")
            print("   • Ambos archivos tienen exactamente los mismos datos")
            print("   • No hay duplicados/faltantes/incompletos según el criterio seleccionado")
            print("\n   Intenta con un análisis diferente o verifica los archivos.")
            wb.close()
            return None

        # Guardar con optimización para archivos grandes y manejo de errores de permisos
        print_progress("Guardando", 90)

        # Verificar si el archivo ya existe y está bloqueado ANTES de intentar guardar
        if os.path.exists(xlsx_path):
            try:
                # Intentar renombrar el archivo temporalmente para ver si está bloqueado
                test_name = xlsx_path + ".test_lock"
                os.rename(xlsx_path, test_name)
                os.rename(test_name, xlsx_path) # Devolver nombre original
            except (PermissionError, OSError):
                # Si falla, es porque está abierto. Cambiamos el nombre directamente.
                print(f"\n⚠ El archivo de destino '{os.path.basename(xlsx_path)}' está ABIERTO o BLOQUEADO.")
                base, ext = os.path.splitext(xlsx_path)
                timestamp = int(time.time())
                xlsx_path = f"{base}_{timestamp}{ext}"
                print(f"   🔄 Guardando con nuevo nombre para evitar errores: {os.path.basename(xlsx_path)}")

        try:
            wb.save(xlsx_path)
        except (PermissionError, OSError) as e:
            # Fallback de último recurso si aún así falla (ej: carpeta sin permisos)
            print(f"\n⚠ Error persistente al guardar en '{os.path.basename(xlsx_path)}': {e}")

            # Intentar en carpeta temporal o con nombre garantizado único
            import random
            timestamp = int(time.time())
            rand = random.randint(1000, 9999)
            base = os.path.splitext(os.path.basename(xlsx_path))[0].split('_')[0] # Quedarse con parte base
            xlsx_path = os.path.join(output_dir, f"{base}_AUTOSAVE_{timestamp}_{rand}.xlsx")

            print(f"   🚨 INTENTO FINAL: Guardando como '{os.path.basename(xlsx_path)}'...")
            try:
                wb.save(xlsx_path)
            except OSError as final_e:
                print("\n❌ ERROR FATAL: No se pudo guardar el reporte en ninguna ubicación.")
                print(f"   Detalle: {final_e}")
                wb.close()
                return None

        wb.close()
        print_progress("Guardando", 100)
        print()  # salto de línea tras la barra

        # Mostrar tamaño del archivo generado y ruta completa
        output_size_mb = os.path.getsize(xlsx_path) / (1024 * 1024)
        ruta_completa = os.path.abspath(xlsx_path)

        print("\n✅ Archivo de reporte guardado exitosamente")
        print(f"   📂 Ruta: {ruta_completa}")
        print(f"   📦 Tamaño: {output_size_mb:.2f} MB")

        if output_size_mb > 10:
            print("   ℹ Archivo grande generado. Puede tardar en abrir en Excel.")

        # Preguntar si desea abrir el archivo
        print("\n" + "="*70)
        respuesta = input("¿Deseas abrir el archivo ahora? (Y/N): ").strip().upper()

        if respuesta == 'Y' or respuesta == 'S' or respuesta == 'YES' or respuesta == 'SI' or respuesta == 'SÍ':
            abrir_archivo_xlsx(ruta_completa)
        else:
            print("\n📋 Puedes abrir el archivo manualmente desde la ruta indicada.")

        return ruta_completa
    except Exception as e:  # pylint: disable=W0718
        print("\n❌ ERROR: No se pudo guardar el archivo Excel.")
        print(f"\n   Razón técnica: {str(e)}")
        print("\n   Posibles causas:")
        print("   • El archivo está abierto en Excel (ciérralo e intenta nuevamente)")
        print("   • No tienes permisos para escribir en la carpeta")
        print("   • El disco está lleno")
        print("   • Hay un problema con las columnas de datos (caracteres especiales)")
        print("\n   Verifica estos puntos e intenta nuevamente.")
        return None


def imprimir_tabla_bonita(df, titulo=None, _max_col_width=50):
    """
    Imprime un DataFrame de manera legible tipo tabla SQL.
    """
    if titulo:
        print(f"\n🔹 {titulo}")

    if df.empty:
        print("   (Tabla vacía)")
        return

    try:
        from tabulate import tabulate
        df_print = df.copy()

        # Limpiar datos para visualización
        df_print = df_print.fillna('')
        for col in df_print.columns:
            # Truncar textos muy largos (ajustado a 30)
            if df_print[col].dtype == 'object':
                df_print[col] = df_print[col].astype(str).str.slice(0, 30)

        print(tabulate(df_print, headers='keys', tablefmt='grid', showindex=False))

    except ImportError:
        # Fallback si no está tabulate
        print("-" * 100)
        print(df.fillna('').to_string(index=False))
        print("-" * 100)
    except Exception as e:  # pylint: disable=W0718
        print(f"Error al imprimir tabla: {e}")
        print(df)

# ===========================================
# FALTANTES
# ==================================
def analyze_faltantes(df_a: pd.DataFrame, df_b: pd.DataFrame, nombre_a: str, nombre_b: str, reportes_dict: dict):
    """Analiza y reporta registros faltantes entre A y B con marcadores visuales"""
    # Usar sets para comparaciones rápidas en grandes volúmenes
    set_b = set(df_b['__KEY__'].unique())
    set_a = set(df_a['__KEY__'].unique())

    print("  ⏳ Generando índices de comparación...")
    print("  ⏳ Calculando restas (A - B) y (B - A)...")
    
    # NUEVO: Agregar columna de verificación en ambos DataFrames
    df_a_marked = df_a.copy()
    df_b_marked = df_b.copy()
    
    # Marcar en A: ✓ si está en B, ✗ si no está
    df_a_marked['ESTADO_EN_B'] = df_a_marked['__KEY__'].apply(
        lambda x: '✓ Encontrado' if x in set_b else '✗ No encontrado'
    )
    
    # Marcar en B: ✓ si está en A, ✗ si no está
    df_b_marked['ESTADO_EN_A'] = df_b_marked['__KEY__'].apply(
        lambda x: '✓ Encontrado' if x in set_a else '✗ No encontrado'
    )
    
    # Filtrar solo los que faltan
    faltantes_en_b = df_a_marked[~df_a_marked['__KEY__'].isin(set_b)]
    faltantes_en_a = df_b_marked[~df_b_marked['__KEY__'].isin(set_a)]
    
    solo_en_a = len(faltantes_en_b)
    solo_en_b = len(faltantes_en_a)
    comunes = len(set_a & set_b)

    # Cálculo de porcentajes precisos
    print("\n📊 RESULTADO DE LA COMPARACIÓN:")
    print(f"   ✓ REGISTROS COMUNES: {comunes:,}")
    print("      (Presentes en ambos archivos)")
    print(f"   ✗ (A - B) Están en {nombre_a} pero NO en {nombre_b}: {solo_en_a:,} usuarios")
    print(f"   ✗ (B - A) Están en {nombre_b} pero NO en {nombre_a}: {solo_en_b:,} usuarios")

    # Guardar versiones completas con marcadores para referencia
    reportes_dict[f'{nombre_a}_COMPLETO_CON_MARCADORES'] = df_a_marked.drop(columns=['__KEY__'])
    reportes_dict[f'{nombre_b}_COMPLETO_CON_MARCADORES'] = df_b_marked.drop(columns=['__KEY__'])

    # Primero: Mostrar TODOS - Faltantes consolidados
    if (not faltantes_en_a.empty) or (not faltantes_en_b.empty):
        df_todos_faltantes = pd.concat(
            [faltantes_en_a.drop(columns=['__KEY__']), faltantes_en_b.drop(columns=['__KEY__'])],
            ignore_index=True
        ) if (not faltantes_en_a.empty and not faltantes_en_b.empty) else (
            faltantes_en_b.drop(columns=['__KEY__']) if not faltantes_en_b.empty else faltantes_en_a.drop(columns=['__KEY__'])
        )

        if not df_todos_faltantes.empty:
            reportes_dict['TODOS - Faltantes'] = df_todos_faltantes.copy()

    # Segundo: Faltantes en B (con marcador)
    if not faltantes_en_b.empty:
        falt_b_clean = faltantes_en_b.drop(columns=['__KEY__'])
        reportes_dict[f'Faltantes en {nombre_b}'] = falt_b_clean
        reportes_dict['Faltantes en B'] = falt_b_clean

        # Si hay columna de hoja origen, separar por hoja
        if '__HOJA_ORIGEN__' in falt_b_clean.columns:
            hojas_unicas = falt_b_clean['__HOJA_ORIGEN__'].unique()
            for hoja in hojas_unicas:
                df_hoja = falt_b_clean[falt_b_clean['__HOJA_ORIGEN__'] == hoja].copy()
                if not df_hoja.empty:
                    # Mantener todas las columnas, solo filtrar las filas faltantes
                    reportes_dict[f'Faltantes en {hoja}'] = df_hoja
                    print(f"  ✅ Faltantes en {hoja}       : {len(df_hoja):6} registros ({len(df_hoja)/len(df_b)*100:5.2f}% de {hoja})")
    else:
        reportes_dict['Faltantes en B'] = pd.DataFrame()

    # Tercero: Faltantes en A (con marcador)
    if not faltantes_en_a.empty:
        falt_a_clean = faltantes_en_a.drop(columns=['__KEY__'])
        reportes_dict[f'Faltantes en {nombre_a}'] = falt_a_clean
        reportes_dict['Faltantes en A'] = falt_a_clean

        # Si hay columna de hoja origen, separar por hoja
        if '__HOJA_ORIGEN__' in falt_a_clean.columns:
            hojas_unicas = falt_a_clean['__HOJA_ORIGEN__'].unique()
            for hoja in hojas_unicas:
                df_hoja = falt_a_clean[falt_a_clean['__HOJA_ORIGEN__'] == hoja].copy()
                if not df_hoja.empty:
                    # Mantener todas las columnas, solo filtrar las filas faltantes
                    reportes_dict[f'Faltantes en {hoja}'] = df_hoja
                    print(f"  ✅ Faltantes en {hoja}       : {len(df_hoja):6} registros ({len(df_hoja)/len(df_a)*100:5.2f}% de {hoja})")
    else:
        reportes_dict['Faltantes en A'] = pd.DataFrame()
        
    return faltantes_en_a, faltantes_en_b
# ================================
# FINAL
# ==================================

# ===========================================
# DUPLICADOS
# ==================================
def analyze_duplicados(df_a: pd.DataFrame, df_b: pd.DataFrame, nombre_a: str, nombre_b: str, 
                      key_desc_a: str, key_desc_b: str, key_a: str, key_b: str, reportes_dict: dict):
    """Analiza y reporta registros duplicados en A y B"""
    print("  ⏳ Identificando duplicados...")
    print("\n  🔍 DEBUG - Análisis de duplicados:")
    print(f"     Clave en {nombre_a}: '{key_desc_a}'")
    print(f"     Clave en {nombre_b}: '{key_desc_b}'")
    print(f"     Total valores en {nombre_a}: {df_a['__KEY__'].count():,}")
    print(f"     Total valores únicos en {nombre_a}: {df_a['__KEY__'].nunique():,}")
    print(f"     Total valores en {nombre_b}: {df_b['__KEY__'].count():,}")
    print(f"     Total valores únicos en {nombre_b}: {df_b['__KEY__'].nunique():,}")

    # Detectar duplicados por la columna RUT (key_a para archivo A, key_b para archivo B)
    print(f"\n  ⏳ Buscando duplicados en {nombre_a}...")
    duplicados_a = df_a[df_a['__KEY__'].duplicated(keep=False)].sort_values('__KEY__')
    print(f"  ⏳ Buscando duplicados en {nombre_b}...")
    duplicados_b = df_b[df_b['__KEY__'].duplicated(keep=False)].sort_values('__KEY__')

    print("\n  📊 RESULTADOS:")
    print(f"Duplicados en {nombre_a}: {len(duplicados_a):,} registros")
    if not duplicados_a.empty:
        ruts_duplicados_a = duplicados_a['__KEY__'].value_counts()
        print(f"  → {len(ruts_duplicados_a)} identificadores diferentes con duplicados")
        print(f"  → Máximas repeticiones: {ruts_duplicados_a.max()} veces")
    else:
        print(f"  ⚠️  No se encontraron duplicados en {nombre_a}")

    print(f"\nDuplicados en {nombre_b}: {len(duplicados_b):,} registros")
    if not duplicados_b.empty:
        ruts_duplicados_b = duplicados_b['__KEY__'].value_counts()
        print(f"  → {len(ruts_duplicados_b)} identificadores diferentes con duplicados")
        print(f"  → Máximas repeticiones: {ruts_duplicados_b.max()} veces")
    else:
        print(f"  ⚠️  No se encontraron duplicados en {nombre_b}")

    # Primero: Mostrar TODOS - Duplicados consolidados
    if (not duplicados_a.empty) or (not duplicados_b.empty):
        df_todos_duplicados = pd.concat(
            [duplicados_a.drop(columns=['__KEY__']), duplicados_b.drop(columns=['__KEY__'])],
            ignore_index=True
        ) if (not duplicados_a.empty and not duplicados_b.empty) else (
            duplicados_b.drop(columns=['__KEY__']) if not duplicados_b.empty else duplicados_a.drop(columns=['__KEY__'])
        )

        if not df_todos_duplicados.empty:
            # Si estamos usando nombres combinados, no podemos filtrar por la columna clave original
            if '__KEY__' in df_todos_duplicados.columns:
                col_id = '__KEY__'
            elif key_a in df_todos_duplicados.columns and key_a != '__KEY__':
                col_id = key_a
            elif key_b in df_todos_duplicados.columns and key_b != '__KEY__':
                col_id = key_b
            else:
                col_id = df_todos_duplicados.columns[0]

            try:
                ruts_todos_dup = df_todos_duplicados[col_id].value_counts().sort_values(ascending=False)

                print(f"\n📊 TODOS - Registros Duplicados ({len(df_todos_duplicados):,} registros):")
                print("\n   🔍 Resumen de duplicados (Top 20):")
                df_counts = ruts_todos_dup.head(20).reset_index()
                df_counts.columns = ['IDENTIFICADOR', 'CANTIDAD']
                df_counts['IDENTIFICADOR'] = df_counts['IDENTIFICADOR'].astype(str).str.replace('|', ' ', regex=False)
                print(df_counts.to_string(index=False))
            except KeyError:
                pass

            reportes_dict['TODOS - Duplicados'] = df_todos_duplicados.copy()

    # Segundo: Duplicados en A
    if not duplicados_a.empty:
        ruts_dup_a = duplicados_a['__KEY__'].value_counts().sort_values(ascending=False)
        print(f"\n1️⃣ Duplicados en {nombre_a} ({len(duplicados_a):,} registros | {len(ruts_dup_a)} únicos):")
        print("\n   🔍 Identificadores duplicados (Top 20):")
        df_counts = ruts_dup_a.head(20).reset_index()
        df_counts.columns = ['IDENTIFICADOR', 'CANTIDAD']
        df_counts['IDENTIFICADOR'] = df_counts['IDENTIFICADOR'].astype(str).str.replace('|', ' ', regex=False)
        print(df_counts.to_string(index=False))
        if len(ruts_dup_a) > 20: 
            print(f"   ... y {len(ruts_dup_a) - 20} identificadores más")

        # NUEVO: Extraer solo las diferencias entre duplicados
        dup_a_diferencias = extract_duplicate_differences(duplicados_a, '__KEY__')
        if not dup_a_diferencias.empty:
            reportes_dict[f'Duplicados en {nombre_a}'] = dup_a_diferencias
            reportes_dict['Duplicados en A'] = dup_a_diferencias
            print(f"\n  ✅ Extraídas {len(dup_a_diferencias)} filas con diferencias en duplicados de {nombre_a}")
        else:
            dup_a_clean = duplicados_a.drop(columns=['__KEY__'])
            reportes_dict[f'Duplicados en {nombre_a}'] = dup_a_clean
            reportes_dict['Duplicados en A'] = dup_a_clean
    else:
        reportes_dict[f'Duplicados en {nombre_a}'] = pd.DataFrame()
        reportes_dict['Duplicados en A'] = pd.DataFrame()

    # Tercero: Duplicados en B
    if not duplicados_b.empty:
        ruts_dup_b = duplicados_b['__KEY__'].value_counts().sort_values(ascending=False)
        print(f"\n2️⃣ Duplicados en {nombre_b} ({len(duplicados_b):,} registros | {len(ruts_dup_b)} únicos):")
        print("\n   🔍 Identificadores duplicados (Top 20):")
        df_counts = ruts_dup_b.head(20).reset_index()
        df_counts.columns = ['IDENTIFICADOR', 'CANTIDAD']
        df_counts['IDENTIFICADOR'] = df_counts['IDENTIFICADOR'].astype(str).str.replace('|', ' ', regex=False)
        print(df_counts.to_string(index=False))
        if len(ruts_dup_b) > 20:
            print(f"   ... y {len(ruts_dup_b) - 20} identificadores más")

        print("\n   📋 Primeros 10 registros duplicados:")
        df_show = duplicados_b.drop(columns=['__KEY__']).head(10)
        imprimir_tabla_bonita(df_show, None)

        # NUEVO: Extraer solo las diferencias entre duplicados
        dup_b_diferencias = extract_duplicate_differences(duplicados_b, '__KEY__')
        if not dup_b_diferencias.empty:
            reportes_dict[f'Duplicados en {nombre_b}'] = dup_b_diferencias
            reportes_dict['Duplicados en B'] = dup_b_diferencias
            print(f"\n  ✅ Extraídas {len(dup_b_diferencias)} filas con diferencias en duplicados de {nombre_b}")
        else:
            dup_b_clean = duplicados_b.drop(columns=['__KEY__'])
            reportes_dict[f'Duplicados en {nombre_b}'] = dup_b_clean
            reportes_dict['Duplicados en B'] = dup_b_clean
    else:
        reportes_dict[f'Duplicados en {nombre_b}'] = pd.DataFrame()
        reportes_dict['Duplicados en B'] = pd.DataFrame()
        
    return duplicados_a, duplicados_b
# ================================
# FINAL
# ==================================


# ===========================================
# INCOMPLETOS
# ==================================
def analyze_incompletos(df_a: pd.DataFrame, df_b: pd.DataFrame, nombre_a: str, nombre_b: str, key_a: str, key_b: str, reportes_dict: dict):
    """Analiza y reporta registros incompletos"""
    incompletos_a = mark_incomplete(df_a, exclude_cols=['__KEY__'])
    incompletos_b = mark_incomplete(df_b, exclude_cols=['__KEY__'])

    # Primero: Mostrar TODOS - Incompletos consolidados
    if (not incompletos_a.empty) or (not incompletos_b.empty):
        df_todos_incompletos = pd.concat(
            [incompletos_a.drop(columns=['__KEY__']), incompletos_b.drop(columns=['__KEY__'])],
            ignore_index=True
        ) if (not incompletos_a.empty and not incompletos_b.empty) else (
            incompletos_b.drop(columns=['__KEY__']) if not incompletos_b.empty else incompletos_a.drop(columns=['__KEY__'])
        )

        if not df_todos_incompletos.empty:
            df_show = df_todos_incompletos.head(5)
            df_show = format_dataframe_rut(df_show, key_a)
            imprimir_tabla_bonita(df_show, f"📊 TODOS - Registros Incompletos ({len(df_todos_incompletos):,} registros):")
            reportes_dict['TODOS - Incompletos'] = df_todos_incompletos.copy()

    # Segundo: Incompletos en A
    if not incompletos_a.empty:
        df_show = incompletos_a.drop(columns=['__KEY__']).head(5)
        df_show = format_dataframe_rut(df_show, key_a)
        imprimir_tabla_bonita(df_show, f"1️⃣ Registros incompletos en {nombre_a} ({len(incompletos_a):,}):")
        # Guardar datos completos sin transformación
        reportes_dict[f'Incompletos en {nombre_a}'] = incompletos_a.drop(columns=['__KEY__'])
        reportes_dict['Incompletos en A'] = incompletos_a.drop(columns=['__KEY__'])
    else:
        reportes_dict[f'Incompletos en {nombre_a}'] = pd.DataFrame()
        reportes_dict['Incompletos en A'] = pd.DataFrame()

    # Tercero: Incompletos en B
    if not incompletos_b.empty:
        df_show = incompletos_b.drop(columns=['__KEY__']).head(5)
        df_show = format_dataframe_rut(df_show, key_b)
        imprimir_tabla_bonita(df_show, f"2️⃣ Registros incompletos en {nombre_b} ({len(incompletos_b):,}):")
        # Guardar datos completos sin transformación
        reportes_dict[f'Incompletos en {nombre_b}'] = incompletos_b.drop(columns=['__KEY__'])
        reportes_dict['Incompletos en B'] = incompletos_b.drop(columns=['__KEY__'])
    else:
        reportes_dict[f'Incompletos en {nombre_b}'] = pd.DataFrame()
        reportes_dict['Incompletos en B'] = pd.DataFrame()
        
    return incompletos_a, incompletos_b
# ================================
# FINAL
# ==================================


def main(file_a: str, file_b: str, key: Optional[str] = None, sheet_a: Optional[str] = None,
        sheet_b: Optional[str] = None, tipos_analisis: Optional[list] = None, iden_config: Optional[dict] = None):
    """Función principal de comparación (soporta múltiples hojas)

    Args:
        tipos_analisis: Lista de tipos de análisis a realizar ['duplicados', 'faltantes', 'incompletos']
        ¿Si es None, se realizan todos los análisis
    """

    # Si no se especifican tipos de análisis, hacer todos
    if tipos_analisis is None:
        tipos_analisis = ['duplicados', 'faltantes', 'incompletos']

    # Validar archivos
    if not os.path.exists(file_a) or not os.path.exists(file_b):
        print("❌ Error: Uno o ambos archivos no existen")
        return

    # Verificar tamaño de archivos
    size_a_mb = os.path.getsize(file_a) / (1024 * 1024)
    size_b_mb = os.path.getsize(file_b) / (1024 * 1024)

    # Extraer nombres de archivos (sin ruta, sin extensión)
    nombre_a = os.path.splitext(os.path.basename(file_a))[0]
    nombre_b = os.path.splitext(os.path.basename(file_b))[0]

    print("\n📊 Información de archivos:")
    print(f"  {nombre_a}: {size_a_mb:.2f} MB")
    print(f"  {nombre_b}: {size_b_mb:.2f} MB")

    if size_a_mb > 8 or size_b_mb > 8:
        print("  ⚡ Archivos grandes detectados - modo optimizado activado")

    # Definir directorio de salida al inicio (mismo donde está el script)
    output_dir = '.'

    # Listar hojas disponibles
    print("📋 Hojas disponibles:")
    sheets_a = get_xlsx_sheets(file_a)
    sheets_b = get_xlsx_sheets(file_b)
    print(f"  {nombre_a}: {sheets_a}")
    print(f"  {nombre_b}: {sheets_b}")

    # Verificar si se deben procesar todas las hojas
    process_all_sheets_a = (sheet_a == 'ALL_SHEETS')
    process_all_sheets_b = (sheet_b == 'ALL_SHEETS')

    # Cargar archivos (con soporte para múltiples hojas)
    if process_all_sheets_a:
        print(f"\n📂 Cargando TODAS las hojas de {nombre_a}...")
        sheets_dict_a = load_all_sheets(file_a)
        if not sheets_dict_a:
            print("❌ No se pudieron cargar las hojas del archivo A")
            return
        # Concatenar todas las hojas AGREGANDO columna de origen
        dfs_with_origin = []
        for sheet_name, df in sheets_dict_a.items():
            df_copy = df.copy()
            df_copy['__HOJA_ORIGEN__'] = sheet_name
            dfs_with_origin.append(df_copy)
        df_a = pd.concat(dfs_with_origin, ignore_index=True)
        print(f"  ✓ Total combinado: {len(df_a):,} filas × {len(df_a.columns)} columnas")
        memory_usage = get_memory_usage(df_a)
        print(f"  💾 Memoria utilizada: {memory_usage}")
        # Verificar si necesita advertencia de memoria
        memory_mb = float(memory_usage.replace(' MB', ''))
        show_memory_warning(memory_mb)
    else:
        print(f"\n📂 Cargando {nombre_a}: {file_a}...")
        use_chunks_a = bool(sheets_a and len(sheets_a) > 0)
        df_a = load_table(file_a, sheet_name=sheet_a, use_chunks=use_chunks_a)
        if df_a.empty:
            return
        print(f"  ✓ {len(df_a):,} filas × {len(df_a.columns)} columnas")
        memory_usage = get_memory_usage(df_a)
        print(f"  💾 Memoria utilizada: {memory_usage}")
        memory_mb = float(memory_usage.replace(' MB', ''))
        show_memory_warning(memory_mb)
    print(f"  Columnas: {', '.join(df_a.columns[:5])}{'...' if len(df_a.columns) > 5 else ''}")

    if process_all_sheets_b:
        print(f"\n📂 Cargando TODAS las hojas de {nombre_b}...")
        sheets_dict_b = load_all_sheets(file_b)
        if not sheets_dict_b:
            print("❌ No se pudieron cargar las hojas del archivo B")
            return
        # Concatenar todas las hojas AGREGANDO columna de origen
        dfs_with_origin = []
        for sheet_name, df in sheets_dict_b.items():
            df_copy = df.copy()
            df_copy['__HOJA_ORIGEN__'] = sheet_name
            dfs_with_origin.append(df_copy)
        df_b = pd.concat(dfs_with_origin, ignore_index=True)
        print(f"  ✓ Total combinado: {len(df_b):,} filas × {len(df_b.columns)} columnas")
        memory_usage = get_memory_usage(df_b)
        print(f"  💾 Memoria utilizada: {memory_usage}")
        memory_mb = float(memory_usage.replace(' MB', ''))
        show_memory_warning(memory_mb)
    else:
        print(f"\n📂 Cargando {nombre_b}: {file_b}...")
        use_chunks_b = bool(sheets_b and len(sheets_b) > 0)
        df_b = load_table(file_b, sheet_name=sheet_b, use_chunks=use_chunks_b)
        if df_b.empty:
            return
        print(f"  ✓ {len(df_b):,} filas × {len(df_b.columns)} columnas")
        memory_usage = get_memory_usage(df_b)
        print(f"  💾 Memoria utilizada: {memory_usage}")
        memory_mb = float(memory_usage.replace(' MB', ''))
        show_memory_warning(memory_mb)
    print(f"  Columnas: {', '.join(df_b.columns[:5])}{'...' if len(df_b.columns) > 5 else ''}")

    # Detectar columnas clave (Priorizando Nombre + Apellidos)
    print("\n🔑 Configurando identificación de registros...")

    key_series_a = None
    key_series_b = None
    key_type_a = ""
    key_type_b = ""

    # 1. MODO PERSONALIZADO (si aplica)
    if iden_config and iden_config.get('mode') == 'manual':
        fields = iden_config.get('fields', [])
        print(f"  🔧 Modo Personalizado Activo: {', '.join(fields).upper()}")
        key_series_a, key_type_a, _ = generate_custom_key(df_a, nombre_a, fields)
        key_series_b, key_type_b, _ = generate_custom_key(df_b, nombre_b, fields)

    # 2. MODO AUTOMÁTICO (si no se generó clave arriba)
    if key_series_a is None or key_series_b is None:
        if iden_config and iden_config.get('mode') == 'manual':
            print("  ⚠️ Falló la configuración personalizada en uno o ambos archivos. Intentando automático...")

        print("  ⏳ Buscando columnas de nombres y apellidos (Automático)...")
        key_series_a, key_type_a, _ = generate_person_key(df_a, nombre_a)
        key_series_b, key_type_b, _ = generate_person_key(df_b, nombre_b)

    key_a = "SISTEMA_DETECT"
    key_b = "SISTEMA_DETECT"

    modo_identificacion = ""

    if key_series_a is not None and key_series_b is not None:
        print("\n  ✅ IDENTIFICACIÓN POR NOMBRE EXITOSA")
        print(f"     Modo: {key_type_a} vs {key_type_b}")
        df_a['__KEY__'] = key_series_a
        df_b['__KEY__'] = key_series_b
        # Guardamos descripción pero usamos __KEY__ como columna de operación
        key_desc_a = f"COMBINADA ({key_type_a})"
        key_desc_b = f"COMBINADA ({key_type_b})"
        key_a = '__KEY__'
        key_b = '__KEY__'
        modo_identificacion = "NOMBRE Y APELLIDOS"
    else:
        # Fallback: Usar RUT u otra columna clave
        print("\n  ⚠️ No se pudieron identificar nombres en ambos archivos. Usando método tradicional (RUT/ID).")
        key_a = auto_detect_key_column(df_a, key)
        key_b = auto_detect_key_column(df_b, key)

        # Buscar mejor coincidencia
        key_a, key_b = find_matching_key_columns(df_a, df_b, key_a, key_b)

        print(f"  {nombre_a} → Usando clave: '{key_a}'")
        print(f"  {nombre_b} → Usando clave: '{key_b}'")

        # Normalizar valores clave tradicionales
        df_a['__KEY__'] = clean_string_for_key(df_a[key_a])
        df_b['__KEY__'] = clean_string_for_key(df_b[key_b])

        key_desc_a = key_a
        key_desc_b = key_b
        modo_identificacion = f"COLUMNA INDIVIDUAL ({key_a})"

    # Verificación de unicidad de la clave generada
    uniq_a = df_a['__KEY__'].nunique() / len(df_a) * 100
    uniq_b = df_b['__KEY__'].nunique() / len(df_b) * 100

    print(f"\n  🔍 Calidad de la clave de identificación ({modo_identificacion}):")
    print(f"     {nombre_a}: {uniq_a:.1f}% registros únicos")
    print(f"     {nombre_b}: {uniq_b:.1f}% registros únicos")

    if uniq_a < 80 or uniq_b < 80:
        print("     ⚠️ ADVERTENCIA: Hay muchos nombres repetidos. La comparación podría generar falsos positivos.")

    # Análisis de diferencias (optimizado para grandes volúmenes)
    print(f"\n{'='*60}")
    print("📊 ANÁLISIS COMPARATIVO")
    print(f"   Clave usada: {modo_identificacion}")
    print(f"{'='*60}")
    print(f"Total en {nombre_a}: {len(df_a):,}")
    print(f"Total en {nombre_b}: {len(df_b):,}")

    # Generar reportes
    print(f"\n{'='*60}")
    print("💾 GENERANDO REPORTES")
    print(f"{'='*60}\n")

    # Diccionario para almacenar todos los reportes
    reportes_dict = {}
    # Guardar nombres para títulos dinámicos
    reportes_dict['_NOMBRE_A'] = nombre_a
    reportes_dict['_NOMBRE_B'] = nombre_b
    reportes_dict['_TOTAL_A'] = len(df_a)
    reportes_dict['_TOTAL_B'] = len(df_b)

    # Inicializar variables para diagnóstico por si no se ejecutan los análisis
    faltantes_en_a = pd.DataFrame()
    faltantes_en_b = pd.DataFrame()
    duplicados_a = pd.DataFrame()
    duplicados_b = pd.DataFrame()
    incompletos_a = pd.DataFrame()
    incompletos_b = pd.DataFrame()

    # 1. ANALISIS FALTANTES
    if 'faltantes' in tipos_analisis:
        faltantes_en_a, faltantes_en_b = analyze_faltantes(df_a, df_b, nombre_a, nombre_b, reportes_dict)
    else:
        reportes_dict['Faltantes en B'] = pd.DataFrame()
        reportes_dict['Faltantes en A'] = pd.DataFrame()

    # 2. ANALISIS DUPLICADOS
    if 'duplicados' in tipos_analisis:
        # Nota: La función analyze_duplicados ya maneja el llenado de reportes_dict
        duplicados_a, duplicados_b = analyze_duplicados(df_a, df_b, nombre_a, nombre_b, 
                                                      key_desc_a, key_desc_b, key_a, key_b, reportes_dict)
    else:
        reportes_dict['Duplicados en A'] = pd.DataFrame()
        reportes_dict['Duplicados en B'] = pd.DataFrame()
        reportes_dict[f'Duplicados en {nombre_a}'] = pd.DataFrame()
        reportes_dict[f'Duplicados en {nombre_b}'] = pd.DataFrame()

    # 3. ANALISIS INCOMPLETOS
    if 'incompletos' in tipos_analisis:
        incompletos_a, incompletos_b = analyze_incompletos(df_a, df_b, nombre_a, nombre_b, key_a, key_b, reportes_dict)
    else:
        reportes_dict['Incompletos en A'] = pd.DataFrame()
        reportes_dict['Incompletos en B'] = pd.DataFrame()
        reportes_dict[f'Incompletos en {nombre_a}'] = pd.DataFrame()
        reportes_dict[f'Incompletos en {nombre_b}'] = pd.DataFrame()

    # Crear tablas consolidadas por categoría - ESTO YA ESTÁ HECHO DIRECTAMENTE EN REPORTES_DICT
    # Los datos TODOS ya se guardaron en reportes_dict como 'TODOS - Faltantes', etc.

    # Análisis de datos nulos (nuevas funcionalidades)
    print(f"\n{'='*60}")
    print("📋 ANÁLISIS DE DATOS NULOS/FALTANTES")
    print(f"{'='*60}")

    # Información sobre nulidades en A
    null_info_a = find_null_data_columns(df_a, exclude_cols=['__KEY__', key_a])
    print_null_stats_table(null_info_a, f"{nombre_a} - Análisis de Nulidades")

    # Información sobre nulidades en B
    null_info_b = find_null_data_columns(df_b, exclude_cols=['__KEY__', key_b])
    print_null_stats_table(null_info_b, f"{nombre_b} - Análisis de Nulidades")

    # NO INCLUIMOS "Usuarios con datos faltantes" pues muestra columnas resumidas
    # Los datos faltantes ya están en 'TODOS - Faltantes' y 'Faltantes en A/B' con columnas completas

    # === ANÁLISIS DE PRIORIDAD / DIAGNÓSTICO (Solicitado) ===
    print(f"\n{'='*60}")
    print("🏥 DIAGNÓSTICO PRIORITARIO")
    print(f"{'='*60}")

    diagnostico_rows = []
    umb_prioridad = 85.0

    # Función auxiliar para evaluar y agregar diagnóstico
    def evaluar_diagnostico(df_target, nombre_df, tipo_problema, total_referencia, mensaje_extra=""):
        if df_target.empty:
            return

        cnt = len(df_target)
        pct = (cnt / total_referencia * 100) if total_referencia > 0 else 0.0

        if pct > umb_prioridad:
            print(f"  ⚠️  ALERTA: Alta tasa de {tipo_problema} en {nombre_df} ({pct:.1f}%)")
            obs = (f"El documento tiene una alta variedad de {tipo_problema}, por lo tanto los datos "
                   f"serían muchos (>85%). {mensaje_extra}")

            diagnostico_rows.append({
                'TIPO': tipo_problema.upper(),
                'ARCHIVO': nombre_df,
                'CANTIDAD': cnt,
                'TOTAL REF': total_referencia,
                '% CAMBIO': f"{pct:.1f}%",
                'PRIORIDAD': 'CRÍTICA',
                'OBSERVACIÓN': obs
            })

    # 1. Evaluar Faltantes (Datos en A que no están en B -> Faltantes en B)
    if 'faltantes' in tipos_analisis:
        # Faltantes en B (A - B): Estan en A pero faltan en B
        evaluar_diagnostico(faltantes_en_b, nombre_b, "faltantes (datos no encontrados)", len(df_a),
                           f"Datos presentes en {nombre_a} pero ausentes en {nombre_b}.")

        # Faltantes en A (B - A): Sobran en B (faltantes en A)
        evaluar_diagnostico(faltantes_en_a, nombre_a, "datos sobrantes (no en base)", len(df_b),
                           f"Datos presentes en {nombre_b} pero ausentes en {nombre_a}.")

    # 2. Evaluar Duplicados
    if 'duplicados' in tipos_analisis:
        evaluar_diagnostico(duplicados_a, nombre_a, "duplicados", len(df_a))
        evaluar_diagnostico(duplicados_b, nombre_b, "duplicados", len(df_b))

    # 3. Evaluar Incompletos
    if 'incompletos' in tipos_analisis:
        evaluar_diagnostico(incompletos_a, nombre_a, "registros incompletos", len(df_a))
        evaluar_diagnostico(incompletos_b, nombre_b, "registros incompletos", len(df_b))

    if diagnostico_rows:
        df_diag = pd.DataFrame(diagnostico_rows)
        # Reordenar columnas para mejor lectura
        cols_order = ['PRIORIDAD', 'TIPO', 'ARCHIVO', '% CAMBIO', 'CANTIDAD', 'OBSERVACIÓN']
        # Asegurar que existan las columnas
        existing_cols = [c for c in cols_order if c in df_diag.columns] + [c for c in df_diag.columns if c not in cols_order]
        df_diag = df_diag[existing_cols]

        reportes_dict['DIAGNOSTICO_PRIORITARIO'] = df_diag
        imprimir_tabla_bonita(df_diag, "RESUMEN DIAGNÓSTICO PRIORITARIO")
    else:
        print("  ✓ No se detectaron problemas críticos (>85%).")

    # Guardar todos los reportes en un único archivo Excel
    ruta_guardada = save_outputs_single_file(reportes_dict, output_dir, tipos_analisis)

    if not ruta_guardada:
        print("\n❌ No se pudo completar la generación del reporte.")
        print("   Revisa los mensajes de error anteriores para más detalles.")

    input("\nPresione Enter para volver al menú principal...")



def menu_seleccion_archivos() -> list:
    """Selecciona múltiples archivos con opción de agregar más"""
    clear_screen()
    print_header()
    print("📋 SELECCIÓN DE ARCHIVOS (Modo Ventana)")
    print("=" * 70)

    print("\n1️⃣  Abriendo ventana para seleccionar archivos (puedes elegir varios a la vez)...")
    lista_archivos = seleccionar_archivos_ventana_multiple("Selecciona los Archivos a Comparar")

    if len(lista_archivos) < 2:
        print("❌ Debes seleccionar al menos dos archivos.")
        return []

    print("\n✅ Archivos seleccionados:")
    for idx, archivo in enumerate(lista_archivos, 1):
        print(f"  {idx}. {os.path.basename(archivo)}")

    # --- CICLO PARA AGREGAR MÁS ARCHIVOS ---
    while True:
        opcion = input("\n¿Deseas agregar más archivos? (s/n): ").strip().lower()
        if opcion == 's':
            nuevos = seleccionar_archivos_ventana_multiple("Selecciona archivos adicionales")
            if nuevos:
                lista_archivos.extend(nuevos)
                print("\n✅ Lista actualizada:")
                for idx, archivo in enumerate(lista_archivos, 1):
                    print(f"  {idx}. {os.path.basename(archivo)}")
            else:
                print("⚠ No se agregaron archivos.")
        elif opcion == 'n':
            break
        else:
            print("❌ Opción no válida. Responde 's' o 'n'.")

    # --- MOSTRAR RESULTADO FINAL ---
    print("\n" + "=" * 70)
    print("📁 LISTA FINAL DE ARCHIVOS A PROCESAR:")
    print("=" * 70)
    for idx, archivo in enumerate(lista_archivos, 1):
        if os.path.exists(archivo):
            tamaño = os.path.getsize(archivo) / (1024 * 1024)
            print(f"  {idx}. {os.path.basename(archivo)} ({tamaño:.2f} MB) ✓")
        else:
            print(f"  {idx}. {os.path.basename(archivo)} ⚠ NO ENCONTRADO")

    return lista_archivos


def main_multiple(lista_archivos: list, tipos_analisis: Optional[list] = None):
    """
    Compara múltiples archivos de la lista.
    Realiza comparaciones de a pares.

    Args:
        tipos_analisis: Lista de tipos de análisis a realizar ['duplicados', 'faltantes', 'incompletos']
    """
    if len(lista_archivos) < 2:
        print("❌ Se necesitan mínimo 2 archivos")
        return

    # Validar que todos existan
    for archivo in lista_archivos:
        if not os.path.exists(archivo):
            print(f"❌ No encontrado: {archivo}")
            return

    # Comparar de forma secuencial: A vs B, A vs C, A vs D, etc.
    for i in range(len(lista_archivos) - 1):
        file_a = lista_archivos[0]  # Siempre comparar con el primero
        file_b = lista_archivos[i + 1]

        print(f"\n{'='*70}")
        print(f"📊 COMPARACIÓN {i + 1}/{len(lista_archivos) - 1}")
        print(f"   Archivo A: {os.path.basename(file_a)}")
        print(f"   Archivo B: {os.path.basename(file_b)}")
        print(f"{'='*70}")

        try:
            main(file_a, file_b, key=None, sheet_a=None, sheet_b=None, tipos_analisis=tipos_analisis)
        except Exception as e:  # pylint: disable=W0718
            print(f"\n❌ ERROR EN COMPARACIÓN {i + 1}")
            print(f"{'='*70}")
            print(f"Tipo de error: {type(e).__name__}")
            print(f"Mensaje: {e}")
            print(f"{'='*70}")
            import traceback
            print("\nDetalle completo del error:")
            traceback.print_exc()
            print(f"\n{'='*70}")

            if i < len(lista_archivos) - 2:
                print(f"⚠️ Quedan {len(lista_archivos) - 2 - i} comparaciones pendientes")
                print(f"\n{'='*70}")
                respuesta = input("¿Continuar con las siguientes comparaciones? (Y/N): ").strip().upper()
                if respuesta == 'N' or respuesta == 'NO':
                    print("\n⏹ Deteniendo comparaciones...")
                    return
            else:
                input("\n📌 Presione Enter para continuar...")

    print("\n✅ Todas las comparaciones completadas.")


def run_app():
    while True:
        clear_screen()
        print_header()
        print("COMPARADOR DE ARCHIVOS - MENÚ PRINCIPAL")
        print("=" * 70)
        print("  1. Seleccion archivo individual (1 archivo vs otro)")
        print("  2. Seleccion archivo interactiva (2 archivos)")
        print("  3. Seleccion archivo múltiple (3+ archivos)")
        print("  x. Detener programa")

        opcion_menu = input("\nEscribe tu opción (1, 2, 3, x): ").strip().lower()
        if opcion_menu == 'x':
            print("\n👋 Saliendo del programa...")
            sys.exit()

        if opcion_menu == "1" or opcion_menu == "2":
            if opcion_menu == "1":
                result = interactive_menu_individual_selection()
            else:
                result = interactive_menu()

            if result[0] == "":
                continue

            file_a, file_b, key, sheet_a, sheet_b, tipos_analisis_main, iden_config = result

            clear_screen()
            print_header()
            print("⏳ Procesando comparación...\n")

            try:
                main(file_a, file_b, key, sheet_a, sheet_b, tipos_analisis_main, iden_config)
            except Exception as e:  # pylint: disable=W0718
                print(f"\n\n{'='*70}")
                print("❌ ❌ ❌ ERROR DURANTE LA COMPARACIÓN ❌ ❌ ❌")
                print(f"{'='*70}")
                print(f"Tipo de error: {type(e).__name__}")
                print(f"Mensaje: {e}")
                print(f"{'='*70}")
                import traceback
                print("\n📋 Detalle completo del error:")
                print("="*70)
                traceback.print_exc()
                print(f"\n{'='*70}")
                print("⚠️  IMPORTANTE: Revisa el error anterior")
                print(f"{'='*70}")
                print("\nPuedes scrollear hacia arriba para ver más detalles del error")
                print(f"{'='*70}")
                input("\n📌 Presione Enter cuando haya terminado de leer el error...")

                while True:
                    print(f"\n{'='*70}")
                    respuesta_menu = input("¿Deseas volver al menú principal? (Y/N): ").strip().upper()
                    if respuesta_menu in ['Y', 'YES', 'S', 'SÍ']:
                        break
                    elif respuesta_menu in ['N', 'NO']:
                        print("\nPuedes seguir viendo el error. Presiona Ctrl+C para salir si lo deseas.")
                        input("📌 Presione Enter para ver el menú de opciones...")
                        print(f"\n{'='*70}")
                        print("Opciones:")
                        print("  Y - Volver al menú principal")
                        print("  N - No volver")
                        print(f"{'='*70}")
                    else:
                        print("Por favor responde Y (Sí) o N (No)")

        elif opcion_menu == "3":
            # Selección múltiple con ventanas
            lista_main = menu_seleccion_archivos()

            if lista_main and len(lista_main) >= 2:
                # Por defecto seleccionamos todos, el filtrado se hace al final
                tipos_analisis_main = None

                clear_screen()
                print_header()
                print("⏳ Procesando comparaciones múltiples...\n")

                try:
                    main_multiple(lista_main, tipos_analisis_main)
                except Exception as e:  # pylint: disable=W0718
                    print(f"\n\n{'='*70}")
                    print("❌ ❌ ❌ ERROR DURANTE LAS COMPARACIONES MÚLTIPLES ❌ ❌ ❌")
                    print(f"{'='*70}")
                    print(f"Tipo de error: {type(e).__name__}")
                    print(f"Mensaje: {e}")
                    print(f"{'='*70}")
                    import traceback
                    print("\n📋 Detalle completo del error:")
                    print("="*70)
                    traceback.print_exc()
                    print(f"\n{'='*70}")
                    print("⚠️  IMPORTANTE: Revisa el error anterior")
                    print(f"{'='*70}")
                    print("\nPuedes scrollear hacia arriba para ver más detalles del error")
                    print(f"{'='*70}")
                    input("\n📌 Presione Enter cuando haya terminado de leer el error...")

                    while True:
                        print(f"\n{'='*70}")
                        respuesta_menu = input("¿Deseas volver al menú principal? (Y/N): ").strip().upper()
                        if respuesta_menu in ['Y', 'YES', 'S', 'SÍ']:
                            break
                        elif respuesta_menu in ['N', 'NO']:
                            print("\nPuedes seguir viendo el error. Presiona Ctrl+C para salir si lo deseas.")
                            input("📌 Presione Enter para ver el menú de opciones...")
                            print(f"\n{'='*70}")
                            print("Opciones:")
                            print("  Y - Volver al menú principal")
                            print("  N - No volver")
                            print(f"{'='*70}")
                        else:
                            print("Por favor responde Y (Sí) o N (No)")
            else:
                print("❌ Operación cancelada.")
                input("\nPresione Enter para volver al menú principal...")

        elif opcion_menu == "4":
            clear_screen()
            print_header()
            print("👋 ¡Hasta luego!")
            break


        else:
            print(f"\n{'='*70}")
            print(f"❌ Opción no válida: '{opcion_menu}'")
            print("Por favor, selecciona 1, 2, 3 o 4")
            print(f"{'='*70}")
            input("\n📌 Presione Enter para continuar...")


if __name__ == '__main__':
    run_app()